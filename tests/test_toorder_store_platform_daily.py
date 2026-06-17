from pathlib import Path

import openpyxl
import pandas as pd

from modules.transform.pipelines.sales import DB_Toorder_store_platform_daily as pipeline


DATEDETAIL_PREFIX = "종합보고서_일별상세_매출보고서"


def _make_datedetail_xlsx(
    path: Path,
    *,
    sheet_date: str = "2026-05-11",
    store: str = "Store A",
    platform: str = "배민",
    price: int = 100,
    receipts: int = 1,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_date
    ws.cell(3, 6).value = store
    ws.cell(7, 1).value = platform
    ws.cell(7, 6).value = price
    ws.cell(7, 9).value = receipts
    ws.cell(8, 1).value = "합계"
    wb.save(path)


def test_find_pending_datedetail_file_matches_month_and_ignores_archived(tmp_path):
    keep = tmp_path / "ignore.xlsx"
    raw = tmp_path / f"{DATEDETAIL_PREFIX}_2026-05_raw.xlsx"
    bak = tmp_path / f"{DATEDETAIL_PREFIX}_2026-05_bak_010101.xlsx"
    temp = tmp_path / f"~${DATEDETAIL_PREFIX}_2026-05.xlsx"
    target = tmp_path / f"{DATEDETAIL_PREFIX}_2026-05.xlsx"

    keep.write_text("x", encoding="utf-8")
    _make_datedetail_xlsx(raw)
    _make_datedetail_xlsx(bak)
    _make_datedetail_xlsx(temp)
    _make_datedetail_xlsx(target)

    assert pipeline._find_pending_datedetail_file(tmp_path, "2026-05") == target
    assert pipeline._find_pending_datedetail_file(tmp_path, "2026-06") is None


def test_upsert_parquet_keeps_last_duplicate_key(tmp_path):
    parquet_path = tmp_path / "toorder_store_platform_daily.parquet"
    old_df = pd.DataFrame(
        [
            {"date": "2026-05-11", "store": "Store A", "platform": "배민", "price": 100, "receipts_num": 1},
            {"date": "2026-05-11", "store": "Store B", "platform": "배민", "price": 50, "receipts_num": 1},
        ]
    )
    pipeline._upsert_parquet(old_df, parquet_path)

    new_df = pd.DataFrame(
        [
            {"date": "2026-05-11", "store": "Store A", "platform": "배민", "price": 200, "receipts_num": 2},
            {"date": "2026-05-12", "store": "Store A", "platform": "쿠팡이츠", "price": 300, "receipts_num": 3},
        ]
    )
    pipeline._upsert_parquet(new_df, parquet_path)

    result = pd.read_parquet(parquet_path).sort_values(["date", "store", "platform"]).reset_index(drop=True)

    assert len(result) == 3
    replaced = result[(result["date"] == "2026-05-11") & (result["store"] == "Store A") & (result["platform"] == "배민")]
    assert int(replaced.iloc[0]["price"]) == 200
    assert int(replaced.iloc[0]["receipts_num"]) == 2


def test_run_toorder_store_platform_daily_uses_pending_datedetail_and_writes_parquet(tmp_path, monkeypatch):
    manual_dir = tmp_path / "manual"
    dest_dir = tmp_path / "dest"
    manual_dir.mkdir()

    pending_path = manual_dir / f"{DATEDETAIL_PREFIX}_2026-05.xlsx"
    _make_datedetail_xlsx(pending_path, store="Manual Store", price=111, receipts=11)

    def _unexpected_download(**_kwargs):
        raise AssertionError("pending datedetail workbook should avoid download")

    monkeypatch.setattr(pipeline, "run_crawling_datedetail_months", _unexpected_download)

    parquet_path = Path(
        pipeline.run_toorder_store_platform_daily(
            date_from="2026-05-11",
            date_to="2026-05-11",
            dest_dir=dest_dir,
            manual_dir=manual_dir,
        )
    )

    result = pd.read_parquet(parquet_path).sort_values(["store", "platform"]).reset_index(drop=True)

    assert parquet_path.name == "toorder_store_platform_daily.parquet"
    assert len(result) == 1
    assert result.iloc[0]["store"] == "Manual Store"
    assert int(result.iloc[0]["price"]) == 111
    assert not pending_path.exists()
    assert (manual_dir / f"{DATEDETAIL_PREFIX}_2026-05_raw.xlsx").exists()


def test_cleanup_datedetail_xlsx_deletes_only_datedetail_files(tmp_path):
    datedetail = tmp_path / f"{DATEDETAIL_PREFIX}_2026-05_raw.xlsx"
    datedetail_bak = tmp_path / f"{DATEDETAIL_PREFIX}_2026-05_raw_bak_010101.xlsx"
    other_toorder = tmp_path / "종합보고서_일별매출보고서_2026-05_raw.xlsx"
    keep = tmp_path / "ignore.xlsx"

    for path in (datedetail, datedetail_bak, other_toorder, keep):
        path.write_text("x", encoding="utf-8")

    result = pipeline.cleanup_datedetail_xlsx(tmp_path)

    assert result == "deleted=2"
    assert not datedetail.exists()
    assert not datedetail_bak.exists()
    assert other_toorder.exists()
    assert keep.exists()
