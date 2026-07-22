"""
홀 매장 주간보고 Excel 생성 파이프라인

입력:
  - MART_DB/hall_sales_target/hall_sale_target.csv      (매출 실적)
  - MART_DB/hall_sales_target/hall_marketing_target.csv (마케팅 실적)
출력:
  - MART_DB/hall_sales_target/hall_weekly_report_{YYMMDD}.xlsx  (날짜별 누적)

레이아웃 (1시트 "주간보고"):
  [표1] 매출 현황 — 월목표 / 월누계 / 월달성률 / 주목표 / 이번주 / 주달성률
  [표2] 마케팅 현황 — 동일 구조
  [표3] AI 진단 요약 (Ollama gpt-oss)
"""

import calendar
import logging
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.transform.utility.paths import MART_DB

logger = logging.getLogger(__name__)

SALE_CSV  = MART_DB / "hall_sales_target" / "hall_sale_target.csv"
MKT_CSV   = MART_DB / "hall_sales_target" / "hall_marketing_target.csv"
XLSX_DIR  = MART_DB / "hall_sales_target"
LLM_LOG_MD = XLSX_DIR / "llm_log.md"

_MKT_COLS = ["플레이스_유입", "홍보물_배포", "쿠폰_회수건수", "인스타_노출", "당근_노출", "네이버_오더"]
_SALE_CHECK_COLS = ["매출", "영수건수", "점심_매출", "점심_영수건수", "저녁_매출", "저녁_영수건수"]
_ALERT_TO = "puding83@kakao.com"
_ALERT_CC = "a17019@kakao.com"

# Windows 경로 (로그 안내용)
_WIN_BASE = str(XLSX_DIR)

# ── 스타일 상수 ────────────────────────────────────────────────
_DARK_FILL  = PatternFill("solid", fgColor="1F3864")  # 제목: 진남색
_HEAD_FILL  = PatternFill("solid", fgColor="2F5597")  # 컬럼 헤더: 진청
_SEC_FILL   = PatternFill("solid", fgColor="17375E")  # 섹션 소제목: 중청
_AI_FILL    = PatternFill("solid", fgColor="EBF3FB")  # AI 진단 배경: 연청

_GREEN  = PatternFill("solid", fgColor="C6EFCE")
_YELLOW = PatternFill("solid", fgColor="FFEB9C")
_RED    = PatternFill("solid", fgColor="FFC7CE")

_WHITE_BOLD  = Font(color="FFFFFF", bold=True, size=11)
_TITLE_FONT  = Font(color="FFFFFF", bold=True, size=13)
_BOLD        = Font(bold=True, size=10)
_NORMAL      = Font(size=10)
_AI_LABEL    = Font(bold=True, size=10, color="1F3864")

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")
_WRAP   = Alignment(horizontal="left",   vertical="top", wrap_text=True)

_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_NUM = "#,##0"
_FMT_KRW = '#,##0"원"'
_FMT_PCT = "0.0%"


# ── 헬퍼 ───────────────────────────────────────────────────────

def _rate(actual, target):
    return actual / target if target else None


def _fill_for(rate):
    if rate is None:
        return None
    return _GREEN if rate >= 1.0 else (_YELLOW if rate >= 0.8 else _RED)


def _cell(ws, row, col, value, font=None, fill=None, align=None,
          fmt=None, border=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font  or _NORMAL
    c.alignment = align or _CENTER
    if fill:   c.fill = fill
    if fmt:    c.number_format = fmt
    if border: c.border = _BORDER
    return c


def _header_row(ws, row, labels, col_start=1):
    for i, lbl in enumerate(labels):
        _cell(ws, row, col_start + i, lbl, font=_WHITE_BOLD,
              fill=_HEAD_FILL, align=_CENTER)


def _data_row(ws, row, label,
              m_tgt, m_act, w_tgt, w_act,
              num_fmt=_FMT_NUM, show_m=True, show_w=True):
    m_rate = _rate(m_act, m_tgt) if show_m else None
    w_rate = _rate(w_act, w_tgt) if show_w else None

    _cell(ws, row, 1, label,  font=_BOLD, align=_LEFT)
    _cell(ws, row, 2, m_tgt,  fmt=num_fmt)
    _cell(ws, row, 3, m_act,  fmt=num_fmt)
    if m_rate is None:
        _cell(ws, row, 4, "—", align=_CENTER)
    else:
        _cell(ws, row, 4, m_rate, fmt=_FMT_PCT, fill=_fill_for(m_rate))
    _cell(ws, row, 5, w_tgt, fmt=num_fmt)
    _cell(ws, row, 6, w_act, fmt=num_fmt)
    if w_rate is None:
        _cell(ws, row, 7, "—", align=_CENTER)
    else:
        _cell(ws, row, 7, w_rate, fmt=_FMT_PCT, fill=_fill_for(w_rate))


def _section(ws, row, title):
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value=f"▶ {title}")
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = _SEC_FILL
    c.alignment = _LEFT


def _save_workbook_replace(wb: Workbook, path, retries: int = 5):
    import time

    tmp_path = path.with_name(f".{path.stem}.tmp{path.suffix}")
    tmp_path.unlink(missing_ok=True)
    wb.save(tmp_path)
    tmp_path.chmod(0o666)
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            if path.exists():
                path.chmod(0o666)
                path.unlink()
            tmp_path.replace(path)
            path.chmod(0o666)
            return path
        except PermissionError as exc:
            last_exc = exc
            if attempt >= retries:
                break
            wait_sec = min(10, attempt * 2)
            logger.warning("Excel 파일 덮어쓰기 재시도: %s (wait=%ss)", path, wait_sec)
            time.sleep(wait_sec)

    raise PermissionError(f"파일이 열려 있거나 OneDrive에서 잠겨 있어 덮어쓸 수 없습니다: {path}") from last_exc


# ── 데이터 로드 ────────────────────────────────────────────────

def _load_sale():
    df = pd.read_csv(SALE_CSV, dtype=str)
    df["입력날짜"] = pd.to_datetime(df["입력날짜"], errors="coerce")
    for c in df.columns:
        if c not in ("입력날짜", "기준주", "기준월"):
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df


def _to_number(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip()
    return pd.to_numeric(cleaned, errors="coerce").fillna(0)


def _load_mkt():
    df = pd.read_csv(MKT_CSV, dtype=str)
    df = df.rename(columns={"인스타_노출_traget": "인스타_노출_target"})
    df["입력날짜"] = pd.to_datetime(df["입력날짜"], errors="coerce")
    df["기준일자"] = pd.to_datetime(df["기준일자"], errors="coerce")
    for c in df.columns:
        if c not in ("입력날짜", "기준일자"):
            df[c] = _to_number(df[c])
    # 기준일자가 비어있으면 입력날짜로 fallback
    df["_기준일"] = df["기준일자"].fillna(df["입력날짜"])
    df["_month"] = df["_기준일"].dt.month
    df["_year"]  = df["_기준일"].dt.year
    return df


def _mkt_weekly_targets(ym, week_start, mkt_monthly):
    cnt = sum(
        1 for i in range(7)
        if (week_start + timedelta(days=i)).strftime("%Y-%m") == ym
    )
    t   = mkt_monthly.get(ym, {})
    dim = calendar.monthrange(int(ym[:4]), int(ym[5:]))[1]
    return {k: round(v / dim * cnt) for k, v in t.items()}


# ── Ollama 진단 ────────────────────────────────────────────────

def _ai_diagnose(sale_ctx: str, mkt_ctx: str):
    try:
        from modules.transform.utility.qwen_client import query_qwen

        sys_prompt = (
            "F&B 매장 분석가. 반드시 한국어로만 답변. 2문장 이내, 숫자 포함, "
            "개선점 1가지만. 번호/기호 없이 자연스럽게."
        )
        sale_diag = query_qwen(
            f"아래 매출 데이터를 2문장으로 요약하세요. 달성률과 가장 취약한 지표, 개선점 1가지.\n\n{sale_ctx}",
            system_prompt=sys_prompt,
        )
        mkt_diag = query_qwen(
            f"아래 마케팅 데이터를 2문장으로 요약하세요. 가장 낮은 채널과 가장 높은 채널, 개선점 1가지.\n\n{mkt_ctx}",
            system_prompt=sys_prompt,
        )
    except Exception as e:
        logger.warning("Ollama 진단 실패 (fallback 사용): %s", e)
        sale_diag = "⚠ AI 진단 불가 — Ollama 연결 또는 모델 실행 실패. 수동 분석이 필요합니다."
        mkt_diag  = "⚠ AI 진단 불가 — Ollama 연결 또는 모델 실행 실패. 수동 분석이 필요합니다."
    return sale_diag, mkt_diag


def append_weekly_ai_log(log_date: str | None = None) -> str:
    """최신 주간보고의 AI 진단을 날짜별 Markdown 로그에 누적 저장한다."""
    report_path = XLSX_DIR / "hall_weekly_report.xlsx"
    if not report_path.exists():
        raise FileNotFoundError(f"hall_weekly_report.xlsx 없음: {report_path}")

    wb = load_workbook(report_path, data_only=True)
    ws = wb.active

    title = str(ws["A1"].value or "").replace("◆", "").strip()
    period = str(ws["A2"].value or "").replace("기준:", "").strip()
    diagnostics = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        label = str(row[0] or "").strip()
        if label in ("매출 진단", "마케팅 진단"):
            diagnostics[label] = str(row[1] or "").strip()

    if log_date is None:
        log_date = date.today().strftime("%Y-%m-%d")

    section_key = f"{log_date} | {title} | {period}"
    marker = f"<!-- hall-ai-log:{section_key} -->"
    entry = (
        f"{marker}\n"
        f"## {section_key}\n\n"
        f"### 매출 진단\n{diagnostics.get('매출 진단', '')}\n\n"
        f"### 마케팅 진단\n{diagnostics.get('마케팅 진단', '')}\n\n"
    )

    if LLM_LOG_MD.exists():
        content = LLM_LOG_MD.read_text(encoding="utf-8")
    else:
        content = "# 홀 주간보고 AI 진단 로그\n\n"

    start = content.find(marker)
    if start >= 0:
        next_start = content.find("<!-- hall-ai-log:", start + len(marker))
        if next_start >= 0:
            content = content[:start] + entry + content[next_start:]
        else:
            content = content[:start] + entry
    else:
        if not content.endswith("\n"):
            content += "\n"
        content += "\n" + entry

    LLM_LOG_MD.parent.mkdir(parents=True, exist_ok=True)
    LLM_LOG_MD.write_text(content, encoding="utf-8")
    logger.info("AI 진단 로그 저장 완료: %s", LLM_LOG_MD)
    return f"완료: {LLM_LOG_MD}"


def _build_missing_html(missing: list, csv_path: str) -> str:
    rows_html = "\n".join(
        f"<tr><td style='padding:6px 12px;border:1px solid #ddd'>{label}</td>"
        f"<td style='padding:6px 12px;border:1px solid #ddd;color:#c0392b'>{reason}</td></tr>"
        for label, reason in missing
    )
    return f"""<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:'Malgun Gothic',Arial,sans-serif;margin:24px;line-height:1.7">
<div style="border-left:4px solid #e74c3c;padding:16px;background:#fdf3f3;border-radius:4px">
  <h2 style="margin:0 0 8px;color:#c0392b">홀 실적 데이터 미입력 알림</h2>
  <p>아래 항목의 마케팅 또는 OKPOS 실적 데이터가 입력되지 않았습니다. 확인 후 다시 실행해주세요.</p>
  <table style="border-collapse:collapse;margin:12px 0">
    <tr style="background:#c0392b;color:#fff">
      <th style="padding:6px 12px;border:1px solid #ddd">항목</th>
      <th style="padding:6px 12px;border:1px solid #ddd">사유</th>
    </tr>
    {rows_html}
  </table>
  <p style="font-size:13px;color:#555">파일 경로: <code>{csv_path}</code></p>
  <p style="font-size:13px;color:#555">
    입력 항목: 플레이스_유입 / 홍보물_배포 / 쿠폰_회수건수 / 인스타_노출 / 당근_노출 / 네이버_오더
  </p>
</div>
<p style="color:#aaa;font-size:11px;margin-top:20px">본 메일은 Airflow DAG에서 자동 발송되었습니다.</p>
</body>
</html>"""


def _build_complete_html(check_dates: list, csv_path: str) -> str:
    start_date = check_dates[0].strftime("%Y-%m-%d")
    end_date = check_dates[-1].strftime("%Y-%m-%d")
    return f"""<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:'Malgun Gothic',Arial,sans-serif;margin:24px;line-height:1.7">
<div style="border-left:4px solid #27ae60;padding:16px;background:#f3fbf6;border-radius:4px">
  <h2 style="margin:0 0 8px;color:#1e8449">홀 실적 데이터 입력 확인 완료</h2>
  <p>최근 7일 마케팅 데이터와 OKPOS 주간 실적이 문제없이 입력되었습니다.</p>
  <p style="font-size:13px;color:#555">마케팅 확인 기간: {start_date} ~ {end_date}</p>
  <p style="font-size:13px;color:#555">파일 경로: <code>{csv_path}</code></p>
</div>
<p style="color:#aaa;font-size:11px;margin-top:20px">본 메일은 Airflow DAG에서 자동 발송되었습니다.</p>
</body>
</html>"""


def _check_sale_input(today: date) -> list:
    if not SALE_CSV.exists():
        logger.warning("hall_sale_target.csv 없음: %s", SALE_CSV)
        return [("OKPOS 주간 실적", "hall_sale_target.csv 파일 없음")]

    df = pd.read_csv(SALE_CSV, dtype=str)
    if "입력날짜" not in df.columns:
        return [("OKPOS 주간 실적", "입력날짜 컬럼 없음")]

    df["입력날짜"] = pd.to_datetime(df["입력날짜"], errors="coerce")
    target_date = pd.Timestamp(today).date()
    row = df[df["입력날짜"].dt.date == target_date].copy()
    if row.empty:
        return [("OKPOS 주간 실적", f"{target_date.strftime('%Y-%m-%d')} 행 없음")]

    for col in _SALE_CHECK_COLS:
        if col in row.columns:
            row[col] = _to_number(row[col])

    total = sum(row[col].sum() for col in _SALE_CHECK_COLS if col in row.columns)
    if total == 0:
        return [("OKPOS 주간 실적", "매출/영수건수 전체 0")]

    return []


def check_marketing_input(**context) -> None:
    """마케팅 최근 7일과 OKPOS 주간 실적 미입력 여부를 HTML 이메일로 알린다."""
    del context

    from modules.transform.utility.mailer import send_email

    today = date.today()
    check_dates = sorted([today - timedelta(days=i) for i in range(1, 8)])

    if not MKT_CSV.exists():
        logger.warning("hall_marketing_target.csv 없음: %s", MKT_CSV)
        missing = [(d.strftime("%Y-%m-%d (%a)"), "마케팅 CSV 파일 없음") for d in check_dates]
        missing.extend(_check_sale_input(today))
        html = _build_missing_html(missing, str(MKT_CSV))
        send_email(
            subject="[홀 실적 데이터 미입력] 확인 요청",
            html_content=html,
            to_emails=_ALERT_TO,
            cc_emails=_ALERT_CC,
        )
        return

    df = pd.read_csv(MKT_CSV, dtype=str)
    if "기준일자" in df.columns:
        df["_기준일"] = pd.to_datetime(df["기준일자"], errors="coerce")
    else:
        df["_기준일"] = pd.Series(pd.NaT, index=df.index)
    if "입력날짜" in df.columns:
        df["_기준일"] = df["_기준일"].fillna(pd.to_datetime(df["입력날짜"], errors="coerce"))

    for col in _MKT_COLS:
        if col in df.columns:
            df[col] = _to_number(df[col])

    entered = set(df["_기준일"].dt.date.dropna())
    missing = []
    for check_date in check_dates:
        if check_date not in entered:
            missing.append((check_date.strftime("%Y-%m-%d (%a)"), "마케팅 행 없음"))
            continue

        row = df[df["_기준일"].dt.date == check_date]
        total = sum(row[col].sum() for col in _MKT_COLS if col in row.columns)
        if total == 0:
            missing.append((check_date.strftime("%Y-%m-%d (%a)"), "마케팅 전체 항목 0"))

    missing.extend(_check_sale_input(today))

    if not missing:
        logger.info("마케팅/OKPOS 실적 입력 확인 완료 - 미입력 항목 없음")
        html = _build_complete_html(check_dates, str(MKT_CSV))
        send_email(
            subject="[홀 실적 데이터 입력 완료] 마케팅/OKPOS 입력 문제 없음",
            html_content=html,
            to_emails=_ALERT_CC,
        )
        return

    missing_dates = []
    for label, _ in missing:
        first = label.split()[0]
        missing_dates.append(first[5:] if len(first) >= 10 and first[4] == "-" else first)
    missing_dates_str = ", ".join(missing_dates)
    html = _build_missing_html(missing, str(MKT_CSV))
    send_email(
        subject=f"[홀 실적 데이터 미입력] {missing_dates_str} 확인 요청",
        html_content=html,
        to_emails=_ALERT_TO,
        cc_emails=_ALERT_CC,
    )
    logger.info("마케팅/OKPOS 실적 미입력 알림 발송: %s", missing_dates_str)


# ── 메인 함수 ──────────────────────────────────────────────────

def build_weekly_report_excel(monthly_targets: dict,
                               marketing_monthly_targets: dict) -> str:
    if not SALE_CSV.exists():
        raise FileNotFoundError(f"hall_sale_target.csv 없음: {SALE_CSV}")
    if not MKT_CSV.exists():
        raise FileNotFoundError(f"hall_marketing_target.csv 없음: {MKT_CSV}")

    sale_df = _load_sale()
    mkt_df  = _load_mkt()

    # ── 최신 주 결정 ──────────────────────────────────────────
    latest    = sale_df.sort_values("입력날짜").iloc[-1]
    ym        = latest["기준월"]          # "2026-05"
    wk_label  = latest["기준주"]          # "5월5주차"
    entry_dt  = latest["입력날짜"]        # 입력날짜(Timestamp)
    week_start = (entry_dt - timedelta(days=7)).date()
    week_end   = (entry_dt - timedelta(days=1)).date()

    yr, mon = int(ym[:4]), int(ym[5:])

    # ── 월 누계 (매출) ────────────────────────────────────────
    mrows = sale_df[sale_df["기준월"] == ym]
    ms = lambda c: int(mrows[c].sum())
    ws = lambda c: int(latest[c])

    mt = monthly_targets.get(ym, {})

    # ── 이번주 targets (매출) — CSV에서 직접 읽음 ─────────────
    wt_sale   = ws("매출_target")
    wt_cnt    = ws("영수건수_target")
    wt_lunch  = ws("점심_매출_target")
    wt_l_cnt  = ws("점심_영수건수_target")
    wt_dinner = ws("저녁_매출_target")
    wt_d_cnt  = ws("저녁_영수건수_target")

    # ── 마케팅 ────────────────────────────────────────────────
    mkt_month  = mkt_df[(mkt_df["_year"] == yr) & (mkt_df["_month"] == mon)]
    # 이번주 = week_start ~ week_end 범위 합산 (기준일자 fallback 기준)
    mkt_week = mkt_df[
        (mkt_df["_기준일"].dt.date >= week_start) &
        (mkt_df["_기준일"].dt.date <= week_end)
    ]
    mm  = lambda c: int(mkt_month[c].sum()) if c in mkt_month.columns else 0
    mw  = lambda c: int(mkt_week[c].sum())  if c in mkt_week.columns and not mkt_week.empty else 0
    mmt = marketing_monthly_targets.get(ym, {})
    mwt = _mkt_weekly_targets(ym, week_start, marketing_monthly_targets)

    # ── AI 진단 ───────────────────────────────────────────────
    def pct(a, t): return f"{a/t*100:.1f}%" if t else "N/A"
    sale_ctx = (
        f"[{wk_label} | {week_start} ~ {week_end}]\n"
        f"- 전체 매출: 이번주 {ws('매출'):,}원 / 목표 {wt_sale:,}원 (달성률 {pct(ws('매출'), wt_sale)})\n"
        f"- 월 누계:   {ms('매출'):,}원 / 월 목표 {mt.get('sale',0):,}원 (달성률 {pct(ms('매출'), mt.get('sale',1))})\n"
        f"- 점심 매출: 이번주 {ws('점심_매출'):,}원 / 목표 {wt_lunch:,}원 (달성률 {pct(ws('점심_매출'), wt_lunch)})\n"
        f"- 저녁 매출: 이번주 {ws('저녁_매출'):,}원 / 목표 {wt_dinner:,}원 (달성률 {pct(ws('저녁_매출'), wt_dinner)})\n"
        f"- 영수건수:  이번주 {ws('영수건수')}건 / 목표 {wt_cnt}건"
    )
    mkt_ctx = (
        f"[{wk_label}]\n"
        f"- 플레이스 유입: {mw('플레이스_유입')}명 / 목표 {mwt.get('플레이스_유입',0)}명\n"
        f"- 홍보물 배포:  {mw('홍보물_배포')}건 / 목표 {mwt.get('홍보물_배포',0)}건\n"
        f"- 쿠폰 회수:   {mw('쿠폰_회수건수')}건 / 목표 {mwt.get('쿠폰_회수',0)}건\n"
        f"- 인스타 노출:  {mw('인스타_노출')}회 / 목표 {mwt.get('인스타_노출',0)}회\n"
        f"- 당근 노출:   {mw('당근_노출')}회 / 목표 {mwt.get('당근_노출',0)}회\n"
        f"- 네이버 오더: {mw('네이버_오더')}건 / 목표 {mwt.get('네이버_오더',0)}건"
    )
    sale_diag, mkt_diag = _ai_diagnose(sale_ctx, mkt_ctx)

    # ── Excel 생성 ────────────────────────────────────────────
    wb = Workbook()
    ws_ = wb.active
    ws_.title = "주간보고"

    col_widths = [18, 15, 15, 11, 15, 15, 11]
    for i, w in enumerate(col_widths, 1):
        ws_.column_dimensions[get_column_letter(i)].width = w

    # 제목
    ws_.merge_cells("A1:G1")
    c = ws_["A1"]
    c.value = f"◆ 홀 주간보고  {wk_label}"
    c.font  = _TITLE_FONT
    c.fill  = _DARK_FILL
    c.alignment = _CENTER
    ws_.row_dimensions[1].height = 30

    ws_.merge_cells("A2:G2")
    c2 = ws_["A2"]
    c2.value = f"기준:  {week_start.strftime('%Y-%m-%d')} ~ {week_end.strftime('%Y-%m-%d')}"
    c2.font  = Font(size=10, italic=True)
    c2.alignment = _CENTER
    ws_.row_dimensions[2].height = 16

    HEADERS = ["구분", "월목표", "월누계", "월달성률", "주목표", "이번주", "주달성률"]

    # ── 표1: 매출 현황 ─────────────────────────────────────────
    _section(ws_, 4, "매출 현황")
    _header_row(ws_, 5, HEADERS)
    ws_.row_dimensions[5].height = 18

    def _ms_aov(prefix_매출, prefix_cnt):
        """월 누계 테이블단가 = 누계매출/누계영수건수"""
        total = ms(prefix_매출)
        cnt   = ms(prefix_cnt)
        return int(total / cnt) if cnt else 0

    sale_rows = [
        # (label, m_tgt, m_act, w_tgt, w_act, num_fmt, show_m, show_w)
        ("전체 매출",    mt.get("sale",0),          ms("매출"),          wt_sale,   ws("매출"),         _FMT_KRW, True,  True),
        ("영수건수",     mt.get("orders",0),          ms("영수건수"),     wt_cnt,    ws("영수건수"),     _FMT_NUM, True,  True),
        ("테이블단가",   mt.get("aov",0),             _ms_aov("매출","영수건수"), mt.get("aov",0), ws("테이블_객단가"), _FMT_KRW, True,  True),
        ("점심 매출",    mt.get("lunch_sale",0),      ms("점심_매출"),     wt_lunch,  ws("점심_매출"),    _FMT_KRW, True,  True),
        ("점심 영수건수",mt.get("lunch_orders",0),    ms("점심_영수건수"), wt_l_cnt,  ws("점심_영수건수"),_FMT_NUM, True,  True),
        ("점심 테이블단가",mt.get("lunch_aov",0),     _ms_aov("점심_매출","점심_영수건수"), mt.get("lunch_aov",0), ws("점심_테이블_객단가"), _FMT_KRW, True,  True),
        ("저녁 매출",    mt.get("dinner_sale",0),     ms("저녁_매출"),     wt_dinner, ws("저녁_매출"),    _FMT_KRW, True,  True),
        ("저녁 영수건수",mt.get("dinner_orders",0),   ms("저녁_영수건수"), wt_d_cnt,  ws("저녁_영수건수"),_FMT_NUM, True,  True),
        ("저녁 테이블단가",mt.get("dinner_aov",0),    _ms_aov("저녁_매출","저녁_영수건수"), mt.get("dinner_aov",0), ws("저녁_테이블_객단가"), _FMT_KRW, True,  True),
    ]
    r = 6
    for lbl, m_t, m_a, w_t, w_a, fmt, sm, sw in sale_rows:
        _data_row(ws_, r, lbl, m_t, m_a, w_t, w_a, num_fmt=fmt, show_m=sm, show_w=sw)
        ws_.row_dimensions[r].height = 17
        r += 1

    r += 1  # 빈 행

    # ── 표2: 마케팅 현황 ───────────────────────────────────────
    _section(ws_, r, "마케팅 현황")
    r += 1
    _header_row(ws_, r, HEADERS)
    ws_.row_dimensions[r].height = 18
    r += 1

    # (표시명, 월누계_csv컬럼, 이번주_csv컬럼, target_key)
    mkt_metrics = [
        ("플레이스 유입", "플레이스_유입", "플레이스_유입", "플레이스_유입"),
        ("홍보물 배포",   "홍보물_배포",   "홍보물_배포",   "홍보물_배포"),
        ("쿠폰 회수",     "쿠폰_회수건수", "쿠폰_회수건수", "쿠폰_회수"),
        ("인스타 노출",   "인스타_노출",   "인스타_노출",   "인스타_노출"),
        ("당근 노출",     "당근_노출",     "당근_노출",     "당근_노출"),
        ("네이버 오더",   "네이버_오더",   "네이버_오더",   "네이버_오더"),
    ]
    for lbl, csv_m, csv_w, tkey in mkt_metrics:
        _data_row(ws_, r, lbl,
                  mmt.get(tkey, 0), mm(csv_m),
                  mwt.get(tkey, 0), mw(csv_w),
                  num_fmt=_FMT_NUM)
        ws_.row_dimensions[r].height = 17
        r += 1

    r += 1  # 빈 행

    # ── 표3: AI 진단 요약 ─────────────────────────────────────
    _section(ws_, r, "AI 진단 요약  (챗Gpt)")
    ws_.row_dimensions[r].height = 18
    r += 1

    for lbl, diag in [("매출 진단", sale_diag), ("마케팅 진단", mkt_diag)]:
        ws_.merge_cells(f"B{r}:G{r}")
        lbl_c = _cell(ws_, r, 1, lbl, font=_AI_LABEL,
                      fill=_AI_FILL, align=_LEFT)
        diag_c = ws_.cell(row=r, column=2, value=diag)
        diag_c.font      = _NORMAL
        diag_c.alignment = _WRAP
        diag_c.fill      = _AI_FILL
        diag_c.border    = _BORDER
        ws_.row_dimensions[r].height = 72
        r += 1

    # ── 저장 ─────────────────────────────────────────────────
    today_str = date.today().strftime("%y%m%d")
    output_path = XLSX_DIR / f"hall_weekly_report_{today_str}.xlsx"
    latest_path = XLSX_DIR / "hall_weekly_report.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path = _save_workbook_replace(wb, output_path)
    _save_workbook_replace(wb, latest_path)   # 고정 파일명 — Excel/PowerBI 참조용

    win_path = str(output_path)
    logger.info("주간보고 저장 완료 (%s)\nWindows: %s", wk_label, win_path)
    return f"완료: {wk_label} → {win_path}"
