"""
홀 매장 일단위 트래킹 Excel 생성 파이프라인

디자인: "일단위 트래킹 (매일 입력)" 이미지 기반
입력:
  - MART_DB/unified_sales_grp/unified_sales_*.parquet  (판매 실적, 자동)
  - MART_DB/hall_sales_target/hall_marketing_target.csv (마케팅, 기준일자 join)
출력:
  - MART_DB/hall_sales_target/hall_daily_report.xlsx  (고정명, 월별 시트)
  - MART_DB/hall_sales_target/hall_daily_report.csv   (고정명, 최신 CSV)
  - MART_DB/hall_sales_target/hall_daily_report_6m.xlsx 등 월별 파일

구조:
  [제목] 일단위 트래킹 (매일 입력)
  [헤더] 날짜 | 점심 영수건수 | 점심 테이블단가 | 저녁 영수건수 | 저녁 테이블단가
         | 플레이스 유입수 | 홍보물 배포(누적) | 쿠폰 회수 | 인스타 노출
         | 당근 노출 | 네이버오더 건수 | 일 매출(자동계산)
  [목표행] 일 목표치 (주황색)
  [데이터] 당월 전체 날짜 × 일별 실적
"""

import calendar
import logging
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.transform.pipelines.db.DB_Hall_Sales_Target import classify_hall_time_slots
from modules.transform.pipelines.db.DB_UnifiedSales_common import iter_unified_sales_files
from modules.transform.utility.paths import MART_DB

logger = logging.getLogger(__name__)

STORE_NAME   = "송파삼전점"
UNIFIED_ROOT = MART_DB / "unified_sales_grp"
MKT_CSV      = MART_DB / "hall_sales_target" / "hall_marketing_target.csv"
OUTPUT_XLSX  = MART_DB / "hall_sales_target" / "hall_daily_report.xlsx"
OUTPUT_CSV   = MART_DB / "hall_sales_target" / "hall_daily_report.csv"

# ── 스타일 ─────────────────────────────────────────────────────
_NAVY    = "1F3864"
_ORANGE  = "F4A460"
_RED_TXT = "C00000"
_GRAY_BG = "F5F5F5"
_WHITE   = "FFFFFF"

_H_FILL  = PatternFill("solid", fgColor=_NAVY)
_T_FILL  = PatternFill("solid", fgColor=_ORANGE)
_G_FILL  = PatternFill("solid", fgColor=_GRAY_BG)

_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_TITLE_FONT  = Font(bold=True, size=14, color=_NAVY)
_SUB_FONT    = Font(size=9,   color="808080")
_H_FONT      = Font(bold=True, size=10, color=_WHITE)
_T_FONT      = Font(bold=True, size=10, color=_WHITE)
_DATE_FONT   = Font(bold=True, size=10)
_NORM_FONT   = Font(size=10)
_SALE_FONT   = Font(size=10, color=_RED_TXT)
_SALE_T_FONT = Font(bold=True, size=10, color=_RED_TXT)

_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left",   vertical="center")
_RGHT = Alignment(horizontal="right",  vertical="center")


# ── 헬퍼 ───────────────────────────────────────────────────────

def _c(ws, row, col, val, font=None, fill=None, align=None, fmt=None, border=True):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font      = font  or _NORM_FONT
    cell.alignment = align or _CTR
    if fill:  cell.fill = fill
    if fmt:   cell.number_format = fmt
    if border: cell.border = _BORDER
    return cell


# ── 데이터 로드 ────────────────────────────────────────────────

def _load_sales(ym: str | None = None) -> pd.DataFrame:
    files = iter_unified_sales_files()
    if not files:
        raise FileNotFoundError(f"unified_sales parquet 없음: {UNIFIED_ROOT}")
    df = pd.concat([pd.read_parquet(f) for f in files], ignore_index=True)
    df["sale_date"]   = pd.to_datetime(df["sale_date"], errors="coerce")
    df["total_price"] = pd.to_numeric(df["total_price"], errors="coerce").fillna(0).astype(int)
    df["order_cnt"]   = pd.to_numeric(df["order_cnt"],   errors="coerce").fillna(0).astype(int)
    df["store"]    = df["store"].astype(str).str.strip()
    df["platform"] = df["platform"].fillna("").astype(str).str.strip()
    mask = (
        (df["store"] == STORE_NAME) &
        (df["platform"] == "홀")
    )
    if ym:
        mask &= df["sale_date"].dt.strftime("%Y-%m") == ym
    df = df[mask].copy()
    return classify_hall_time_slots(df)  # 빈 df도 time_slot 컬럼 추가됨


def _to_number(series: pd.Series) -> pd.Series:
    cleaned = series.fillna("").astype(str).str.replace(",", "", regex=False).str.strip()
    return pd.to_numeric(cleaned, errors="coerce").fillna(0).astype(int)


def _load_marketing(ym: str) -> pd.DataFrame:
    if not MKT_CSV.exists():
        return pd.DataFrame()
    df = pd.read_csv(MKT_CSV, dtype=str)
    df = df.rename(columns={"인스타_노출_traget": "인스타_노출_target"})
    df["입력날짜"] = pd.to_datetime(df["입력날짜"], errors="coerce")
    df["기준일자"] = pd.to_datetime(df["기준일자"], errors="coerce")
    df["_기준일"] = df["기준일자"].fillna(df["입력날짜"])
    for col in ["플레이스_유입", "홍보물_배포", "쿠폰_회수건수", "인스타_노출", "당근_노출", "네이버_오더"]:
        if col in df.columns:
            df[col] = _to_number(df[col])
    df = df[df["_기준일"].dt.strftime("%Y-%m") == ym]
    df = df.sort_values("_기준일").reset_index(drop=True)
    if "홍보물_배포" in df.columns:
        df["홍보물_배포_누적"] = df["홍보물_배포"].cumsum().astype(int)
    return df


# ── 일별 집계 ──────────────────────────────────────────────────

def _daily_rows(sales_df: pd.DataFrame, mkt_df: pd.DataFrame, all_dates: list) -> list:
    rows = []
    mkt_by_date = {}
    if not mkt_df.empty:
        for _, r in mkt_df.iterrows():
            if not pd.isnull(r["_기준일"]):
                mkt_by_date[r["_기준일"].date()] = r

    for d in all_dates:
        day_s  = sales_df[sales_df["sale_date"].dt.date == d]
        lunch  = day_s[day_s["time_slot"] == "점심"]
        dinner = day_s[day_s["time_slot"] == "저녁"]

        l_cnt  = int(lunch["order_cnt"].sum())
        l_sale = int(lunch["total_price"].sum())
        l_aov  = l_sale // l_cnt if l_cnt else 0
        d_cnt  = int(dinner["order_cnt"].sum())
        d_sale = int(dinner["total_price"].sum())
        d_aov  = d_sale // d_cnt if d_cnt else 0
        tot    = l_sale + d_sale

        m = mkt_by_date.get(d)
        rows.append({
            "date":        d,
            "lunch_cnt":   l_cnt  or None,
            "lunch_sale":  l_sale or None,
            "lunch_aov":   l_aov  or None,
            "dinner_cnt":  d_cnt  or None,
            "dinner_sale": d_sale or None,
            "dinner_aov":  d_aov  or None,
            "place":       int(m["플레이스_유입"])    if m is not None else None,
            "leaflet_cum": int(m["홍보물_배포_누적"]) if m is not None and "홍보물_배포_누적" in m.index else None,
            "coupon":      int(m["쿠폰_회수건수"])      if m is not None else None,
            "insta":       int(m["인스타_노출"])      if m is not None else None,
            "karrot":      int(m["당근_노출"])        if m is not None else None,
            "naver":       int(m["네이버_오더"])      if m is not None else None,
            "daily_sale":  tot,
        })
    return rows


# ── Excel 작성 ─────────────────────────────────────────────────

HEADERS = [
    "날짜",
    "점심\n영수건수",
    "점심\n매출",
    "점심\n테이블단가",
    "저녁\n영수건수",
    "저녁\n매출",
    "저녁\n테이블단가",
    "플레이스\n유입수",
    "홍보물\n배포(누적)",
    "쿠폰\n회수",
    "인스타\n노출",
    "당근\n노출",
    "네이버오더\n건수",
    "일\n매출",
]
N_COLS = len(HEADERS)

COL_WIDTHS = [7, 8, 12, 11, 8, 12, 11, 9, 11, 7, 8, 8, 10, 12]


def _write_sheet(ws, rows: list, daily_target: dict) -> None:
    # 열 너비
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 제목 (row 1) ──────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    title = ws["A1"]
    title.value     = "일단위 트래킹 (매일 입력)"
    title.font      = _TITLE_FONT
    title.alignment = _LEFT
    ws.row_dimensions[1].height = 26

    # ── 헤더 (row 2) ──────────────────────────────────────────
    for i, h in enumerate(HEADERS, 1):
        _c(ws, 2, i, h, font=_H_FONT, fill=_H_FILL, align=_CTR)
    ws.row_dimensions[2].height = 30

    # ── 목표 (row 3) ──────────────────────────────────────────
    goal = [
        "목표",
        daily_target.get("lunch_orders",  0),
        daily_target.get("lunch_sale",    0),
        daily_target.get("lunch_aov",     0),
        daily_target.get("dinner_orders", 0),
        daily_target.get("dinner_sale",   0),
        daily_target.get("dinner_aov",    0),
        daily_target.get("place",         0),
        daily_target.get("홍보물_배포",   0),
        daily_target.get("coupon",        0),
        daily_target.get("insta",         0),
        daily_target.get("karrot",        0),
        daily_target.get("naver",         0),
        daily_target.get("sale",          0),
    ]
    for i, val in enumerate(goal, 1):
        if i == 1:
            _c(ws, 3, i, val, font=_T_FONT, fill=_T_FILL, align=_CTR)
        elif i in (3, 4, 6, 7, 14):
            _c(ws, 3, i, val, font=_T_FONT, fill=_T_FILL, align=_RGHT, fmt="#,##0")
        else:
            _c(ws, 3, i, val, font=_T_FONT, fill=_T_FILL, align=_CTR, fmt="#,##0")
    ws.row_dimensions[3].height = 18

    # ── 데이터 행 ─────────────────────────────────────────────
    for idx, row in enumerate(rows):
        r     = 4 + idx
        is_odd = idx % 2 == 0   # 0-indexed → row 4,6,8... = 흰색
        bg    = None if is_odd else _G_FILL
        d     = row["date"]
        label = f"{d.month}/{d.day}"

        def dc(col, val, fmt=None, is_sale=False):
            font  = (_SALE_FONT if is_sale else _NORM_FONT)
            align = (_RGHT if is_sale or (fmt and "0" in fmt) else _CTR)
            _c(ws, r, col, val, font=font, fill=bg, align=align, fmt=fmt, border=True)

        # 날짜
        _c(ws, r, 1, label, font=_DATE_FONT, fill=bg, align=_CTR)

        # 점심 영수건수, 테이블단가
        dc(2,  row["lunch_cnt"],  "#,##0")
        dc(3,  row["lunch_sale"], "#,##0", is_sale=True)
        dc(4,  row["lunch_aov"],  "#,##0")
        # 저녁 영수건수, 테이블단가
        dc(5,  row["dinner_cnt"], "#,##0")
        dc(6,  row["dinner_sale"], "#,##0", is_sale=True)
        dc(7,  row["dinner_aov"], "#,##0")
        # 마케팅
        dc(8,  row["place"],        "#,##0")
        dc(9,  row["leaflet_cum"],  "#,##0")
        dc(10, row["coupon"],       "#,##0")
        dc(11, row["insta"],        "#,##0")
        dc(12, row["karrot"],       "#,##0")
        dc(13, row["naver"],        "#,##0")
        dc(14, row["daily_sale"],   "#,##0", is_sale=True)

        ws.row_dimensions[r].height = 17

def _save_workbook_replace(wb: Workbook, path: Path) -> None:
    tmp_path = path.with_name(f".{path.stem}.tmp{path.suffix}")
    tmp_path.unlink(missing_ok=True)
    wb.save(tmp_path)
    tmp_path.chmod(0o666)
    if path.exists():
        try:
            path.chmod(0o666)
            path.unlink()
        except PermissionError as e:
            raise PermissionError(f"파일이 열려 있거나 OneDrive에서 잠겨 있습니다: {path}") from e
    tmp_path.replace(path)
    path.chmod(0o666)


def _month_file_path(ym: str) -> Path:
    month = int(ym[5:7])
    return OUTPUT_XLSX.with_name(f"hall_daily_report_{month}m.xlsx")


def _month_dates(ym: str) -> list:
    year, month = int(ym[:4]), int(ym[5:7])
    days_in_month = calendar.monthrange(year, month)[1]
    return [date(year, month, d) for d in range(1, days_in_month + 1)]


def _available_months(sales_df: pd.DataFrame) -> list[str]:
    months = set()
    if not sales_df.empty:
        months.update(sales_df["sale_date"].dt.strftime("%Y-%m").dropna().unique())
    months.add(date.today().strftime("%Y-%m"))
    return sorted(months, reverse=True)


def _build_month_rows(sales_df: pd.DataFrame, ym: str) -> list:
    month_sales = sales_df[sales_df["sale_date"].dt.strftime("%Y-%m") == ym].copy()
    mkt_df = _load_marketing(ym)
    return _daily_rows(month_sales, mkt_df, _month_dates(ym))


def _write_excel_files(sales_df: pd.DataFrame, months: list[str], daily_target: dict) -> None:
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    del wb[wb.active.title]

    for ym in months:
        rows = _build_month_rows(sales_df, ym)
        ws = wb.create_sheet(title=ym)
        _write_sheet(ws, rows, daily_target)

        month_wb = Workbook()
        month_ws = month_wb.active
        month_ws.title = ym
        _write_sheet(month_ws, rows, daily_target)
        _save_workbook_replace(month_wb, _month_file_path(ym))

    _save_workbook_replace(wb, OUTPUT_XLSX)


def _safe_div(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0


def _int0(value) -> int:
    if pd.isna(value):
        return 0
    return int(value)


def _month_target(ym: str, monthly_targets: dict, daily_target: dict, key: str) -> int:
    if ym in monthly_targets:
        return int(monthly_targets[ym].get(key, 0))
    year, month = int(ym[:4]), int(ym[5:7])
    return int(daily_target.get(key, 0) * calendar.monthrange(year, month)[1])


def _marketing_month_target(ym: str, marketing_monthly_targets: dict, key: str) -> int:
    return int(marketing_monthly_targets.get(ym, {}).get(key, 0))


def _forecast_value(actual: pd.Series, current_day: int, fallback: int = 0) -> int:
    values = pd.to_numeric(actual, errors="coerce").fillna(0)
    observed = values[values != 0]
    if observed.empty:
        return int(fallback)
    return int(round(observed.sum() / max(current_day, len(observed))))


def _monthly_csv_rows(
    sales_df: pd.DataFrame,
    ym: str,
    monthly_targets: dict,
    marketing_monthly_targets: dict,
    daily_target: dict,
) -> list[dict]:
    rows = _build_month_rows(sales_df, ym)
    if not rows:
        return []

    df = pd.DataFrame(rows)
    df["sale_date"] = pd.to_datetime(df["date"])
    df["day"] = df["sale_date"].dt.day

    max_actual = sales_df["sale_date"].max() if not sales_df.empty else pd.NaT
    current_ym = date.today().strftime("%Y-%m")
    if pd.isna(max_actual):
        actual_cutoff = pd.Timestamp.min
    elif ym < current_ym:
        actual_cutoff = df["sale_date"].max()
    else:
        actual_cutoff = min(pd.Timestamp(max_actual), df["sale_date"].max())

    actual_mask = df["sale_date"] <= actual_cutoff
    current_day = int(actual_mask.sum()) if actual_mask.any() else 0

    forecast = {
        "daily_sale": _forecast_value(df.loc[actual_mask, "daily_sale"], current_day, daily_target.get("sale", 0)),
        "lunch_sale": _forecast_value(df.loc[actual_mask, "lunch_sale"], current_day, daily_target.get("lunch_sale", 0)),
        "dinner_sale": _forecast_value(df.loc[actual_mask, "dinner_sale"], current_day, daily_target.get("dinner_sale", 0)),
        "lunch_cnt": _forecast_value(df.loc[actual_mask, "lunch_cnt"], current_day, daily_target.get("lunch_orders", 0)),
        "dinner_cnt": _forecast_value(df.loc[actual_mask, "dinner_cnt"], current_day, daily_target.get("dinner_orders", 0)),
        "place": _forecast_value(df.loc[actual_mask, "place"], current_day, daily_target.get("place", 0)),
        "leaflet_cum": _forecast_value(df.loc[actual_mask, "leaflet_cum"], current_day, daily_target.get("홍보물_배포", 0)),
        "coupon": _forecast_value(df.loc[actual_mask, "coupon"], current_day, daily_target.get("coupon", 0)),
        "insta": _forecast_value(df.loc[actual_mask, "insta"], current_day, daily_target.get("insta", 0)),
        "karrot": _forecast_value(df.loc[actual_mask, "karrot"], current_day, daily_target.get("karrot", 0)),
        "naver": _forecast_value(df.loc[actual_mask, "naver"], current_day, daily_target.get("naver", 0)),
    }

    output = []
    cumul = {
        "sale": 0,
        "orders": 0,
        "lunch_sale": 0,
        "lunch_orders": 0,
        "dinner_sale": 0,
        "dinner_orders": 0,
        "place": 0,
        "leaflet": 0,
        "coupon": 0,
        "insta": 0,
        "karrot": 0,
        "naver": 0,
    }

    month_sale_target = _month_target(ym, monthly_targets, daily_target, "sale")
    month_aov_target = _month_target(ym, monthly_targets, daily_target, "aov")
    month_order_target = _month_target(ym, monthly_targets, daily_target, "orders")
    month_lunch_sale_target = _month_target(ym, monthly_targets, daily_target, "lunch_sale")
    month_dinner_sale_target = _month_target(ym, monthly_targets, daily_target, "dinner_sale")
    month_place_target = _marketing_month_target(ym, marketing_monthly_targets, "플레이스_유입")
    month_leaflet_target = _marketing_month_target(ym, marketing_monthly_targets, "홍보물_배포")
    month_insta_target = _marketing_month_target(ym, marketing_monthly_targets, "인스타_노출")
    month_karrot_target = _marketing_month_target(ym, marketing_monthly_targets, "당근_노출")

    for _, row in df.iterrows():
        is_actual = row["sale_date"] <= actual_cutoff
        lunch_sale = _int0(row["lunch_sale"]) if is_actual else forecast["lunch_sale"]
        dinner_sale = _int0(row["dinner_sale"]) if is_actual else forecast["dinner_sale"]
        total_amt = _int0(row["daily_sale"]) if is_actual else forecast["daily_sale"]
        lunch_cnt = _int0(row["lunch_cnt"]) if is_actual else forecast["lunch_cnt"]
        dinner_cnt = _int0(row["dinner_cnt"]) if is_actual else forecast["dinner_cnt"]
        total_orders = lunch_cnt + dinner_cnt
        place = _int0(row["place"]) if is_actual else forecast["place"]
        leaflet = _int0(row["leaflet_cum"]) if is_actual else forecast["leaflet_cum"]
        coupon = _int0(row["coupon"]) if is_actual else forecast["coupon"]
        insta = _int0(row["insta"]) if is_actual else forecast["insta"]
        karrot = _int0(row["karrot"]) if is_actual else forecast["karrot"]
        naver = _int0(row["naver"]) if is_actual else forecast["naver"]

        cumul["sale"] += total_amt
        cumul["orders"] += total_orders
        cumul["lunch_sale"] += lunch_sale
        cumul["lunch_orders"] += lunch_cnt
        cumul["dinner_sale"] += dinner_sale
        cumul["dinner_orders"] += dinner_cnt
        cumul["place"] += place
        cumul["leaflet"] = max(cumul["leaflet"], leaflet)
        cumul["coupon"] += coupon
        cumul["insta"] += insta
        cumul["karrot"] += karrot
        cumul["naver"] += naver

        output.append({
            "sale_date": row["sale_date"].strftime("%Y-%m-%d"),
            "ym": ym.replace("-", "_"),
            "구분": "실제값" if is_actual else "예측값",
            "total_amt": total_amt,
            "월매출_목표": month_sale_target,
            "누적매출": cumul["sale"],
            "tot_order_cnt": total_orders,
            "일영수건수_목표": month_order_target,
            "누적주문수": cumul["orders"],
            "일영수건수(일평균)": _safe_div(cumul["orders"], row["day"]),
            "객단가": _safe_div(total_amt, total_orders),
            "객단가_목표": month_aov_target,
            "누적_객단가": _safe_div(cumul["sale"], cumul["orders"]),
            "점심_매출": lunch_sale,
            "점심_일매출목표": month_lunch_sale_target,
            "누적_점심_매출": cumul["lunch_sale"],
            "점심_주문건수": lunch_cnt,
            "누적_점심_주문건수": cumul["lunch_orders"],
            "점심_객단가": _safe_div(lunch_sale, lunch_cnt),
            "누적_점심_객단가": _safe_div(cumul["lunch_sale"], cumul["lunch_orders"]),
            "저녁_매출": dinner_sale,
            "저녁_일매출목표": month_dinner_sale_target,
            "누적_저녁_매출": cumul["dinner_sale"],
            "저녁_주문건수": dinner_cnt,
            "누적_저녁_주문건수": cumul["dinner_orders"],
            "저녁_객단가": _safe_div(dinner_sale, dinner_cnt),
            "누적_저녁_객단가": _safe_div(cumul["dinner_sale"], cumul["dinner_orders"]),
            "플레이스_유입": place,
            "플레이스_유입_목표": month_place_target,
            "플레이스_유입_누적": cumul["place"],
            "플레이스_유입_일평균": _safe_div(cumul["place"], row["day"]),
            "플레이스_유입_일평균_목표": daily_target.get("place", 0),
            "홍보물_배포": leaflet,
            "홍보물_배포_목표": month_leaflet_target,
            "홍보물_배포_누적": cumul["leaflet"],
            "쿠폰_회수건수": coupon,
            "쿠폰_회수율_목표": 0.7,
            "쿠폰_회수건수_누적": cumul["coupon"],
            "쿠폰_회수율": 0.7,
            "인스타_노출": insta,
            "인스타_노출_목표": month_insta_target,
            "인스타_노출_누적": cumul["insta"],
            "당근_노출": karrot,
            "당근_노출_목표": month_karrot_target,
            "당근_노출_누적": cumul["karrot"],
            "네이버_오더": naver,
            "네이버_오더_일평균_목표": daily_target.get("naver", 0),
            "네이버_오더_누적": cumul["naver"],
            "네이버_오더_일평균": _safe_div(cumul["naver"], row["day"]),
            "점심_영수건수_목표": daily_target.get("lunch_orders", 0) * row["day"],
            "점심_테이블단가_목표": daily_target.get("lunch_aov", 0),
            "점심_일매출_일평균_목표": daily_target.get("lunch_sale", 0),
            "저녁_영수건수_목표": daily_target.get("dinner_orders", 0) * row["day"],
            "저녁_테이블단가_목표": daily_target.get("dinner_aov", 0),
            "저녁_일매출_일평균_목표": daily_target.get("dinner_sale", 0),
        })

    return output


def build_daily_tracking_csv(
    monthly_targets: dict,
    marketing_monthly_targets: dict,
    daily_target: dict,
) -> str:
    sales_df = _load_sales()
    months = _available_months(sales_df)
    rows = []
    for ym in sorted(months):
        rows.extend(
            _monthly_csv_rows(
                sales_df,
                ym,
                monthly_targets,
                marketing_monthly_targets,
                daily_target,
            )
        )

    OUTPUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(OUTPUT_CSV, index=False, encoding="utf-8-sig")
    logger.info("일단위 트래킹 CSV 저장 완료: %d행 → %s", len(rows), OUTPUT_CSV)
    return f"완료 {len(rows)}행 → {OUTPUT_CSV}"


# ── 메인 함수 ──────────────────────────────────────────────────

def build_daily_tracking_excel(monthly_targets: dict,
                                marketing_monthly_targets: dict,
                                daily_target: dict) -> str:
    sales_df = _load_sales()
    months = _available_months(sales_df)
    _write_excel_files(sales_df, months, daily_target)

    win_path = str(OUTPUT_XLSX)
    logger.info("일단위 트래킹 저장 완료 (%s)\nWindows: %s",
                ", ".join(months), win_path)
    return f"완료: {', '.join(months)} → {win_path}"
