from __future__ import annotations

import csv
import os
import time
from html import escape
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple
import json

import openpyxl
import pendulum
import pandas as pd

from modules.transform.utility.paths import ANALYTICS_DB


AI_DAILY_COLLECTION_DIR = ANALYTICS_DB / "ai_daily_collection"
INTEGRATED_XLSX_FILENAME = "종합보고서_일별매출보고서_통합.xlsx"
DAILY_SUMMARY_CSV_FILENAME = "종합보고서_일별매출보고서_일별합계.csv"
INTEGRATED_SHEET_NAME = "Sheet1"
DAILY_SUMMARY_SHEET_NAME = "일별합계"
PARQUET_FILENAME = "toorder_daily_store_all.parquet"


def _default_ai_daily_collection_dir() -> Path:
    return AI_DAILY_COLLECTION_DIR


EXCLUDED_STORE_NAMES = {
    "합계",
    "테스트매장",
    "테스트 매장",
}


def _safe_int(value) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(value)
    try:
        return int(str(value).replace(",", "").strip())
    except Exception:
        return 0


def _find_daily_sheet_2(wb: "openpyxl.Workbook") -> "openpyxl.worksheet.worksheet.Worksheet":
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.cell(3, 1).value == "날짜":
            return ws
    return wb[wb.sheetnames[-1]]


def _extract_report_dates(value) -> Tuple[Optional[str], Optional[str]]:
    if not isinstance(value, str):
        return None, None

    text = value.strip()
    if not text:
        return None, None

    parts = [part.strip() for part in text.split("~")]
    if len(parts) == 2 and all(part[:4].isdigit() and "-" in part for part in parts):
        return parts[0], parts[1]
    if text[:4].isdigit() and "-" in text:
        return text, text
    return None, None


def _is_excluded_store(store: str) -> bool:
    return (not store) or store in EXCLUDED_STORE_NAMES or "합계" in store


def _parse_legacy_daily_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
) -> Tuple[str, List[Tuple[str, str, int, int]]]:
    stores: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if not isinstance(v, str):
            continue
        store = v.strip()
        if store == "날짜" or _is_excluded_store(store):
            continue
        stores[store] = c

    rows: List[Tuple[str, str, int, int]] = []
    first_date = last_date = None
    for r in range(7, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if not isinstance(v, str) or not v[:4].isdigit() or "-" not in v:
            continue
        date_str = v.strip()
        if first_date is None:
            first_date = date_str
        last_date = date_str
        for store, col in stores.items():
            sales = _safe_int(ws.cell(r, col).value)
            order_count = _safe_int(ws.cell(r, col + 3).value)
            rows.append((date_str, store, order_count, sales))

    if not rows:
        raise ValueError(f"리포트 날짜를 찾지 못했습니다: {ws.title}")

    report_date = first_date if first_date == last_date else f"{first_date}~{last_date}"
    return report_date, rows


def _parse_channel_daily_sheet(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
) -> Tuple[str, List[Tuple[str, str, int, int]]]:
    start_date, end_date = _extract_report_dates(ws.cell(2, 1).value)
    if not start_date:
        raise ValueError(f"리포트 날짜를 찾지 못했습니다: {ws.title}")

    rows: List[Tuple[str, str, int, int]] = []
    for c in range(6, ws.max_column + 1, 4):
        store_value = ws.cell(3, c).value
        if not isinstance(store_value, str):
            continue

        store = store_value.strip()
        if _is_excluded_store(store):
            continue

        sales = 0
        order_count = 0
        for r in range(7, ws.max_row + 1):
            channel = ws.cell(r, 1).value
            if not isinstance(channel, str):
                continue
            channel = channel.strip()
            if not channel or channel == "합계":
                continue
            sales += _safe_int(ws.cell(r, c).value)
            order_count += _safe_int(ws.cell(r, c + 3).value)

        rows.append((start_date, store, order_count, sales))

    if not rows:
        raise ValueError(f"채널별 일별 리포트를 해석하지 못했습니다: {ws.title}")

    report_date = start_date if start_date == end_date else f"{start_date}~{end_date}"
    return report_date, rows


def parse_toorder_daily_report(
    xlsx_path: Path,
) -> Tuple[str, List[Tuple[str, str, int, int]]]:
    """
    투오더 '종합보고서_일별매출보고서' 엑셀에서 날짜/매장별 (주문건수, 매출)을 추출한다.

    Returns:
        (report_date, [(date, store, order_count, sales), ...])
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = _find_daily_sheet_2(wb)

    if ws.cell(3, 1).value == "날짜":
        return _parse_legacy_daily_sheet(ws)
    if ws.cell(3, 1).value == "채널":
        return _parse_channel_daily_sheet(ws)

    first_sheet_title = str(wb[wb.sheetnames[0]].cell(1, 1).value or "")
    if "채널별" in ws.title or "채널별" in first_sheet_title:
        return _parse_channel_daily_sheet(ws)

    return _parse_legacy_daily_sheet(ws)


def ensure_integrated_workbook(path: Path) -> None:
    if path.exists():
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["날짜", "매장", "주문건수", "매출"])
    wb.save(path)


def _normalize_sale_date(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return pd.to_datetime(text).strftime("%Y-%m-%d")
    except Exception:
        return text[:10] if len(text) >= 10 else text


def _normalize_store_name(value) -> str:
    return str(value or "").strip()


def _rows_to_dataframe(rows: Iterable[Tuple[str, str, int, int]]) -> pd.DataFrame:
    records = []
    for sale_date, store, order_count, sales in rows:
        normalized_date = _normalize_sale_date(sale_date)
        normalized_store = _normalize_store_name(store)
        if not normalized_date or _is_excluded_store(normalized_store):
            continue
        records.append(
            {
                "sale_date": normalized_date,
                "store": normalized_store,
                "order_count": _safe_int(order_count),
                "sales": _safe_int(sales),
            }
        )
    return pd.DataFrame(records, columns=["sale_date", "store", "order_count", "sales"])


def _all_parquet_path(base_dir: Path) -> Path:
    return base_dir / PARQUET_FILENAME


def _load_daily_parquet(parquet_path: Path) -> pd.DataFrame:
    df = pd.read_parquet(parquet_path, engine="pyarrow")
    if df.empty:
        return pd.DataFrame(columns=["sale_date", "store", "order_count", "sales"])

    renamed = df.rename(
        columns={
            "date": "sale_date",
            "주문건수": "order_count",
            "매출": "sales",
        }
    ).copy()
    for col in ["sale_date", "store", "order_count", "sales"]:
        if col not in renamed.columns:
            renamed[col] = None

    renamed["sale_date"] = renamed["sale_date"].apply(_normalize_sale_date)
    renamed["store"] = renamed["store"].apply(_normalize_store_name)
    renamed["order_count"] = pd.to_numeric(renamed["order_count"], errors="coerce").fillna(0).astype(int)
    renamed["sales"] = pd.to_numeric(renamed["sales"], errors="coerce").fillna(0).astype(int)
    renamed = renamed.loc[
        renamed["sale_date"].notna()
        & renamed["store"].astype(str).str.strip().ne("")
        & ~renamed["store"].astype(str).apply(_is_excluded_store)
    ]
    return renamed[["sale_date", "store", "order_count", "sales"]].reset_index(drop=True)


def load_ai_daily_collection_detail_df(
    base_dir: Optional[Path] = None,
    target_date: Optional[str] = None,
) -> pd.DataFrame:
    resolved_base_dir = Path(base_dir) if base_dir else _default_ai_daily_collection_dir()
    parquet_path = _all_parquet_path(resolved_base_dir)
    if not parquet_path.exists():
        return pd.DataFrame(columns=["sale_date", "store", "order_count", "sales"])

    df = _load_daily_parquet(parquet_path)
    if target_date:
        df = df.loc[df["sale_date"] == target_date]
    return df.sort_values(["sale_date", "store"]).reset_index(drop=True)


def load_ai_daily_collection_daily_totals(base_dir: Optional[Path] = None) -> pd.DataFrame:
    detail_df = load_ai_daily_collection_detail_df(base_dir=base_dir)
    if detail_df.empty:
        return pd.DataFrame(columns=["sale_date", "order_count", "sales"])

    grouped = (
        detail_df.groupby("sale_date", as_index=False)[["order_count", "sales"]]
        .sum()
        .sort_values("sale_date")
        .reset_index(drop=True)
    )
    grouped["order_count"] = grouped["order_count"].astype(int)
    grouped["sales"] = grouped["sales"].astype(int)
    return grouped


def rebuild_ai_daily_collection_compat_outputs(
    base_dir: Optional[Path] = None,
    integrated_path: Optional[Path] = None,
    csv_path: Optional[Path] = None,
) -> Dict[str, Path]:
    resolved_base_dir = Path(base_dir) if base_dir else _default_ai_daily_collection_dir()
    resolved_integrated_path = Path(integrated_path) if integrated_path else resolved_base_dir / INTEGRATED_XLSX_FILENAME
    resolved_csv_path = Path(csv_path) if csv_path else resolved_base_dir / DAILY_SUMMARY_CSV_FILENAME

    detail_df = load_ai_daily_collection_detail_df(base_dir=resolved_base_dir)
    totals_df = load_ai_daily_collection_daily_totals(base_dir=resolved_base_dir)

    wb = openpyxl.Workbook()
    ws_detail = wb.active
    ws_detail.title = INTEGRATED_SHEET_NAME
    ws_detail.append(["일자", "매장", "주문건수", "매출"])
    for row in detail_df.itertuples(index=False):
        ws_detail.append([row.sale_date, row.store, int(row.order_count), int(row.sales)])

    ws_summary = wb.create_sheet(DAILY_SUMMARY_SHEET_NAME)
    ws_summary.append(["일자", "주문건수", "매출"])
    for row in totals_df.itertuples(index=False):
        ws_summary.append([row.sale_date, int(row.order_count), int(row.sales)])

    _atomic_save(wb, resolved_integrated_path)

    resolved_csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(resolved_csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["일자", "주문건수", "매출"])
        for row in totals_df.itertuples(index=False):
            writer.writerow([row.sale_date, int(row.order_count), int(row.sales)])

    return {
        "integrated_xlsx": resolved_integrated_path,
        "daily_summary_csv": resolved_csv_path,
    }


def _upsert_rows_to_daily_parquet(base_dir: Path, rows_df: pd.DataFrame) -> int:
    if rows_df.empty:
        return 0

    parquet_path = _all_parquet_path(base_dir)
    new_df = rows_df[["sale_date", "store", "order_count", "sales"]].copy()

    if parquet_path.exists():
        existing_df = _load_daily_parquet(parquet_path)
        keep_existing = ~existing_df.set_index(["sale_date", "store"]).index.isin(
            new_df.set_index(["sale_date", "store"]).index
        )
        merged_df = pd.concat(
            [existing_df.loc[keep_existing, ["sale_date", "store", "order_count", "sales"]], new_df],
            ignore_index=True,
        )
    else:
        merged_df = new_df

    merged_df = (
        merged_df.drop_duplicates(subset=["sale_date", "store"], keep="last")
        .sort_values(["sale_date", "store"])
        .reset_index(drop=True)
    )
    tmp_path = parquet_path.with_name(f"{parquet_path.stem}._tmp{parquet_path.suffix}")
    merged_df.to_parquet(tmp_path, index=False, engine="pyarrow")
    os.replace(tmp_path, parquet_path)
    return len(rows_df)


def upsert_integrated_rows(
    integrated_path: Path,
    rows: Iterable[Tuple[str, str, int, int]],
) -> int:
    base_dir = integrated_path.parent
    base_dir.mkdir(parents=True, exist_ok=True)
    rows_df = _rows_to_dataframe(rows)
    changed = _upsert_rows_to_daily_parquet(base_dir, rows_df)
    rebuild_ai_daily_collection_compat_outputs(
        base_dir=base_dir,
        integrated_path=integrated_path,
        csv_path=base_dir / DAILY_SUMMARY_CSV_FILENAME,
    )
    return changed


def _atomic_save(wb: "openpyxl.Workbook", path: Path) -> None:
    tmp = path.with_name(f"{path.stem}._tmp{path.suffix}")
    wb.save(tmp)
    try:
        for _ in range(5):
            try:
                os.replace(tmp, path)
                return
            except PermissionError:
                time.sleep(1.0)
        pending = path.with_name(
            f"{path.stem}_pending_{time.strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        )
        os.replace(tmp, pending)
        raise PermissionError(
            f"통합 파일이 사용 중이라 덮어쓸 수 없습니다: {path} (대신 {pending} 로 저장됨)"
        )
    finally:
        if tmp.exists():
            try:
                tmp.unlink()
            except Exception:
                pass


def build_daily_summary_sheet(
    integrated_path: Path,
    sheet_name: str = "날짜별합계",
) -> int:
    """
    Sheet1(메인) 데이터를 날짜별로 groupby-sum해 별도 시트에 기록한다.
    합계 행은 메인 시트에 추가하지 않는다.

    Returns:
        날짜 수
    """
    ensure_integrated_workbook(integrated_path)
    wb = openpyxl.load_workbook(integrated_path, data_only=True)
    ws_main = wb.active

    totals: Dict[str, Tuple[int, int]] = {}
    for r in range(2, ws_main.max_row + 1):
        d = ws_main.cell(r, 1).value
        s = ws_main.cell(r, 2).value
        if not isinstance(d, str) or not isinstance(s, str):
            continue
        date_str = d.strip()
        store = s.strip()
        if not date_str or not store:
            continue
        if store in EXCLUDED_STORE_NAMES or "합계" in store:
            continue
        oc = _safe_int(ws_main.cell(r, 3).value)
        sales = _safe_int(ws_main.cell(r, 4).value)
        prev_oc, prev_sales = totals.get(date_str, (0, 0))
        totals[date_str] = (prev_oc + oc, prev_sales + sales)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws_summary = wb.create_sheet(sheet_name)
    ws_summary.append(["날짜", "주문건수", "매출"])
    for date_str, (oc, sales) in sorted(totals.items()):
        ws_summary.append([date_str, oc, sales])

    _atomic_save(wb, integrated_path)
    return len(totals)


def export_daily_summary_csv(integrated_path: Path, csv_path: Path) -> int:
    """날짜별합계 시트를 CSV로 내보낸다. 반환값: 데이터 행 수."""
    wb = openpyxl.load_workbook(integrated_path, data_only=True)
    if "날짜별합계" not in wb.sheetnames:
        raise ValueError(f"날짜별합계 시트가 없습니다: {integrated_path}")
    ws = wb["날짜별합계"]
    rows = list(ws.iter_rows(values_only=True))
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(rows)
    return len(rows) - 1  # 헤더 제외


def _is_ai_daily_collection_source_xlsx(path: Path, integrated_path: Optional[Path] = None) -> bool:
    if not path.is_file():
        return False
    if path.suffix.lower() != ".xlsx":
        return False
    if path.name.startswith("~$"):
        return False
    if path.name.endswith("_raw.xlsx"):
        return False
    if "통합" in path.name:
        return False
    if integrated_path and path.resolve() == integrated_path.resolve():
        return False
    return path.name.startswith("종합보고서_일별매출보고서")


def cleanup_ai_daily_collection_source_xlsx(
    base_dir: Optional[Path] = None,
    extra_dirs: Optional[Iterable[Path]] = None,
) -> List[Path]:
    resolved_base_dir = Path(base_dir) if base_dir else _default_ai_daily_collection_dir()
    cleanup_dirs = [resolved_base_dir, *(extra_dirs or [])]
    deleted: List[Path] = []
    seen = set()

    for directory in cleanup_dirs:
        if not directory:
            continue
        directory = Path(directory)
        for pattern in ("종합보고서_일별매출보고서*_raw.xlsx", "종합보고서_일별매출보고서*_raw_bak_*.xlsx"):
            for path in sorted(directory.glob(pattern)):
                key = str(path).lower()
                if key in seen:
                    continue
                seen.add(key)
                if not path.is_file():
                    continue
                if "통합" in path.name:
                    continue
                try:
                    path.unlink()
                    deleted.append(path)
                except FileNotFoundError:
                    continue
    return deleted


def rename_ai_daily_collection_xlsx_to_raw(src_path: Path) -> Path:
    if src_path.exists():
        src_path.unlink()
    return src_path


def ingest_ai_daily_collection_xlsx(
    src_path: Path,
    *,
    base_dir: Optional[Path] = None,
    rename_to_raw: bool = False,
) -> Dict[str, object]:
    resolved_base_dir = Path(base_dir) if base_dir else _default_ai_daily_collection_dir()
    resolved_base_dir.mkdir(parents=True, exist_ok=True)
    report_date, rows = parse_toorder_daily_report(src_path)
    changed = _upsert_rows_to_daily_parquet(resolved_base_dir, _rows_to_dataframe(rows))
    raw_path = rename_ai_daily_collection_xlsx_to_raw(src_path) if (rename_to_raw or src_path.exists()) else src_path
    return {
        "source": src_path,
        "raw_path": raw_path,
        "report_date": report_date,
        "upsert_rows": changed,
    }


def ingest_pending_ai_daily_collection_xlsx_files(
    *,
    base_dir: Optional[Path] = None,
    extra_dirs: Optional[Iterable[Path]] = None,
    rename_to_raw: bool = False,
) -> List[Dict[str, object]]:
    resolved_base_dir = Path(base_dir) if base_dir else _default_ai_daily_collection_dir()
    resolved_base_dir.mkdir(parents=True, exist_ok=True)
    integrated_path = resolved_base_dir / INTEGRATED_XLSX_FILENAME

    candidates: List[Path] = []
    for directory in [resolved_base_dir, *(extra_dirs or [])]:
        if not directory:
            continue
        for path in sorted(Path(directory).glob("종합보고서_일별매출보고서*.xlsx")):
            if _is_ai_daily_collection_source_xlsx(path, integrated_path=integrated_path):
                candidates.append(path)

    seen = set()
    results: List[Dict[str, object]] = []
    for path in candidates:
        key = str(path).lower()
        if key in seen:
            continue
        seen.add(key)
        results.append(
            ingest_ai_daily_collection_xlsx(
                path,
                base_dir=resolved_base_dir,
                rename_to_raw=rename_to_raw,
            )
        )

    cleanup_ai_daily_collection_source_xlsx(base_dir=resolved_base_dir, extra_dirs=extra_dirs)
    return results


def migrate_ai_daily_collection_xlsx_to_parquet(
    *,
    base_dir: Optional[Path] = None,
    rename_to_raw: bool = False,
) -> List[Dict[str, object]]:
    return ingest_pending_ai_daily_collection_xlsx_files(
        base_dir=base_dir,
        rename_to_raw=rename_to_raw,
    )


def build_daily_summary_sheet(
    integrated_path: Path,
    sheet_name: str = DAILY_SUMMARY_SHEET_NAME,
) -> int:
    del sheet_name
    rebuild_ai_daily_collection_compat_outputs(
        base_dir=integrated_path.parent,
        integrated_path=integrated_path,
        csv_path=integrated_path.parent / DAILY_SUMMARY_CSV_FILENAME,
    )
    totals_df = load_ai_daily_collection_daily_totals(base_dir=integrated_path.parent)
    return len(totals_df)


def export_daily_summary_csv(integrated_path: Path, csv_path: Path) -> int:
    rebuild_ai_daily_collection_compat_outputs(
        base_dir=integrated_path.parent,
        integrated_path=integrated_path,
        csv_path=csv_path,
    )
    totals_df = load_ai_daily_collection_daily_totals(base_dir=integrated_path.parent)
    return len(totals_df)


# ---------------------------------------------------------------------------
# AI daily collection - score & diagnosis (rule-based, v1)
# ---------------------------------------------------------------------------


def _parse_ymd(date_str: str) -> pendulum.Date:
    return pendulum.parse(date_str).date()


def _iter_dates(start: pendulum.Date, end: pendulum.Date) -> Iterable[pendulum.Date]:
    d = start
    while d <= end:
        yield d
        d = d.add(days=1)


def _sum_sales_by_range(daily_sales: Dict[str, int], start: pendulum.Date, end: pendulum.Date) -> int:
    total = 0
    for d in _iter_dates(start, end):
        total += int(daily_sales.get(d.format("YYYY-MM-DD"), 0) or 0)
    return total


def _pct_change(curr: int, base: int) -> Optional[float]:
    if base <= 0:
        return None
    return (curr / base - 1.0) * 100.0


def _score_short(pct: Optional[float]) -> int:
    # 단기: 0% 이상 0점 / -5%~-10% 1점 / -10% 이하 2점
    if pct is None:
        return 0
    if pct >= 0:
        return 0
    if pct <= -10:
        return 2
    if pct <= -5:
        return 1
    return 0


def _score_mid(pct: Optional[float]) -> int:
    # 중기: 0% 이상 0점 / -3%~-7% 1점 / -7% 이하 2점
    if pct is None:
        return 0
    if pct >= 0:
        return 0
    if pct <= -7:
        return 2
    if pct <= -3:
        return 1
    return 0


def _score_long(pct: Optional[float]) -> int:
    # 장기: 중기와 동일 점수 구간
    return _score_mid(pct)


def _score_yoy(pct: Optional[float]) -> int:
    # 전년: 단기와 동일 점수 구간
    return _score_short(pct)


def _grade(total_score: int) -> str:
    if total_score <= 0:
        return "정상"
    if 1 <= total_score <= 2:
        return "주의"
    if 3 <= total_score <= 4:
        return "경고"
    return "위험"


def _fmt_pct(pct: Optional[float]) -> str:
    if pct is None:
        return "N/A"
    return f"{pct:+.1f}%"


def compute_ai_sales_scorecard(
    daily_sales: Dict[str, int],
    target_date: str,
) -> Dict[str, object]:
    """
    기획서 v1(룰 기반) 기준으로 전사 매출 점수/등급 입력값을 계산한다.

    daily_sales:
        {"YYYY-MM-DD": sales_int, ...}  (전사 일매출 합계)
    """
    d0 = _parse_ymd(target_date)
    today_sales = int(daily_sales.get(d0.format("YYYY-MM-DD"), 0) or 0)

    # 1) 단기: 어제 vs 최근 4주 같은요일 평균 (데이터 있는 주만 평균)
    weekday = d0.weekday()  # Monday=0
    same_weekday_sales: List[int] = []
    for w in range(1, 5):
        di = d0.subtract(weeks=w)
        if di.weekday() != weekday:
            continue
        v = daily_sales.get(di.format("YYYY-MM-DD"))
        if v is not None:
            same_weekday_sales.append(int(v or 0))
    short_base_avg = int(round(sum(same_weekday_sales) / len(same_weekday_sales))) if same_weekday_sales else 0
    short_pct = _pct_change(today_sales, short_base_avg) if short_base_avg > 0 else None

    # 2) 중기: 최근 14일 누적 vs 이전 14일 누적 (양쪽 모두 14일)
    recent14_start = d0.subtract(days=13)
    recent14_end = d0
    prev14_start = d0.subtract(days=27)
    prev14_end = d0.subtract(days=14)
    mid_recent14 = _sum_sales_by_range(daily_sales, recent14_start, recent14_end)
    mid_prev14 = _sum_sales_by_range(daily_sales, prev14_start, prev14_end)
    mid_pct = _pct_change(mid_recent14, mid_prev14) if mid_prev14 > 0 else None

    # 3) 장기: 이번달 누적(MTD) vs 전월 동일 경과일 누적(MTD)
    mtd_start = d0.start_of("month")
    mtd_curr = _sum_sales_by_range(daily_sales, mtd_start, d0)
    mtd_day_count = d0.day
    mtd_avg = int(round(mtd_curr / mtd_day_count)) if mtd_day_count > 0 else 0
    elapsed_days = d0.day - 1  # 0-based offset from month start
    prev_month_same_day = d0.subtract(months=1)
    prev_mtd_start = prev_month_same_day.start_of("month")
    prev_mtd_end = prev_mtd_start.add(days=elapsed_days)
    long_prev_mtd = _sum_sales_by_range(daily_sales, prev_mtd_start, prev_mtd_end)
    long_pct = _pct_change(mtd_curr, long_prev_mtd) if long_prev_mtd > 0 else None
    prev_month_start = prev_month_same_day.start_of("month")
    prev_month_end = prev_month_start.end_of("month")
    prev_month_total = _sum_sales_by_range(daily_sales, prev_month_start, prev_month_end)
    prev_month_day_count = prev_month_start.days_in_month
    prev_month_avg = int(round(prev_month_total / prev_month_day_count)) if prev_month_day_count > 0 else 0
    mtd_avg_vs_prev_month_avg_pct = _pct_change(mtd_avg, prev_month_avg) if prev_month_avg > 0 else None

    # 4) 전년: 이번달 누적(MTD) vs 전년 동월 동일 경과일 누적(MTD)
    yoy_same_day = d0.subtract(years=1)
    yoy_mtd_start = yoy_same_day.start_of("month")
    yoy_mtd_end = yoy_mtd_start.add(days=elapsed_days)
    yoy_prev_mtd = _sum_sales_by_range(daily_sales, yoy_mtd_start, yoy_mtd_end)
    yoy_pct = _pct_change(mtd_curr, yoy_prev_mtd) if yoy_prev_mtd > 0 else None

    short_score = _score_short(short_pct)
    mid_score = _score_mid(mid_pct)
    long_score = _score_long(long_pct)
    yoy_score = _score_yoy(yoy_pct)
    total_score = short_score + mid_score + long_score + yoy_score
    grade = _grade(total_score)

    return {
        "target_date": target_date,
        "today_sales": today_sales,
        "short_base_avg": short_base_avg,
        "short_pct": short_pct,
        "short_score": short_score,
        "mid_recent14": mid_recent14,
        "mid_prev14": mid_prev14,
        "mid_pct": mid_pct,
        "mid_score": mid_score,
        "mtd_curr": mtd_curr,
        "mtd_day_count": mtd_day_count,
        "mtd_avg": mtd_avg,
        "long_prev_mtd": long_prev_mtd,
        "long_pct": long_pct,
        "long_score": long_score,
        "prev_month_total": prev_month_total,
        "prev_month_day_count": prev_month_day_count,
        "prev_month_avg": prev_month_avg,
        "mtd_avg_vs_prev_month_avg_pct": mtd_avg_vs_prev_month_avg_pct,
        "yoy_prev_mtd": yoy_prev_mtd,
        "yoy_pct": yoy_pct,
        "yoy_score": yoy_score,
        "total_score": total_score,
        "grade": grade,
    }


def build_rule_based_ai_diagnosis(scorecard: Dict[str, object]) -> Dict[str, str]:
    """
    기획서 7.x (룰 기반 문구) 기반 3줄 진단 문구를 생성한다.
    출력은 1줄씩(현상황/예상원인/제안방향)로 제한한다.
    """
    short_score = int(scorecard["short_score"])
    mid_score = int(scorecard["mid_score"])
    long_score = int(scorecard["long_score"])
    yoy_score = int(scorecard["yoy_score"])
    grade = str(scorecard["grade"])

    short_pct = scorecard.get("short_pct")
    mid_pct = scorecard.get("mid_pct")
    long_pct = scorecard.get("long_pct")
    yoy_pct = scorecard.get("yoy_pct")

    # 1) 현상황: 점수>0 인 기준 중 우선 2개만
    items: List[Tuple[int, float, str, Optional[float]]] = []
    for name, sc, pct in [
        ("전일 매출", short_score, short_pct),
        ("최근 14일 누적", mid_score, mid_pct),
        ("전월 동일 구간", long_score, long_pct),
        ("전년 동월 동일 구간", yoy_score, yoy_pct),
    ]:
        if sc <= 0 or pct is None:
            continue
        items.append((sc, float(pct), name, pct))

    items.sort(key=lambda x: (-x[0], x[1]))  # score desc, pct asc(더 하락한 것 먼저)
    if not items:
        situation = f"- 모든 기준이 정상으로 전사 매출 상태는 {grade}입니다."
    elif len(items) == 1:
        _, _, name, pct = items[0]
        situation = f"- {name} {_fmt_pct(pct)}로 전사 매출 상태는 {grade}입니다."
    else:
        _, _, n1, p1 = items[0]
        _, _, n2, p2 = items[1]
        situation = f"- {n1} {_fmt_pct(p1)}, {n2} {_fmt_pct(p2)}로 전사 매출 상태는 {grade}입니다."

    # 2) 예상 원인 + 3) 제안방향: 점수 조합 기반(원인 단정 금지)
    if short_score > 0 and mid_score > 0 and long_score > 0 and yoy_score > 0:
        cause = "- 일시적 변동보다는 전사 매출 체력 약화 가능성이 높아 원인 분해가 필요합니다."
        suggest = "- 하락 매장 TOP5에서 주문수 감소와 매출 감소 기여도를 먼저 확인한 뒤, 객단가·채널별 매출을 추가 분해해 주세요."
    elif short_score > 0 and mid_score > 0:
        cause = "- 단기 급락과 중기 하락이 동시에 발생해 주문수 감소 또는 일부 매장 집중 하락 가능성이 우선 의심됩니다."
        suggest = "- 하락 매장 TOP5에서 주문수 감소와 매출 감소 기여도를 먼저 확인하고, 객단가·채널별 매출을 추가 분해해 주세요."
    elif mid_score > 0 and long_score > 0:
        cause = "- 최근 매출 흐름 둔화가 월 누적 매출 약화로 이어지고 있을 가능성이 있습니다."
        suggest = "- 최근 14일 기준 하락 매장을 우선 확인하고, 주문수 감소 매장과 채널별 하락 여부를 분리해 주세요."
    elif long_score > 0 and yoy_score > 0:
        cause = "- 전월 대비뿐 아니라 전년 대비 매출 체력도 약화되었을 가능성이 있습니다."
        suggest = "- 전년 대비 반복 하락 매장 여부와 운영 매장 수 변화를 먼저 확인해 주세요."
    elif short_score > 0:
        cause = "- 일시적 매출 변동 또는 특정 매장 급락 가능성이 있습니다."
        suggest = "- 당일 급락 매장에서 휴무/운영시간/주문 오류/배달앱 노출 이슈를 우선 확인해 주세요."
    elif mid_score > 0:
        cause = "- 전사적 하락보다는 최근 2주 흐름 둔화 영향일 가능성이 있습니다."
        suggest = "- 최근 14일 기준 주문수 감소 매장과 채널별 매출 감소 여부를 우선 확인해 주세요."
    elif long_score > 0:
        cause = "- 월 누적 매출이 약화되고 있을 가능성이 있습니다."
        suggest = "- 이번달 목표 달성률과 객단가/포장·배달 비중 변화를 점검해 주세요."
    elif yoy_score > 0:
        cause = "- 전년 동월 대비 매출 체력이 약화되었을 가능성이 있습니다."
        suggest = "- 전년 대비 동일 매장 매출과 시즌 변수(프로모션/운영)를 우선 점검해 주세요."
    else:
        cause = "- 현재 제공된 기준에서 뚜렷한 하락 신호는 없습니다."
        suggest = "- 동일 기준이 반복되는지 추이를 모니터링해 주세요."

    return {"situation": situation, "cause": cause, "suggest": suggest}


_OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://host.docker.internal:11434")
_GPT_MODEL_CANDIDATES = [
    "gpt-oss:20b",
    "gpt-oss:latest",
    "gpt-oss",
    "qwen2.5:7b",
    "qwen2.5:latest",
]


def _get_ollama_client_for_diagnosis():
    import ollama

    client = ollama.Client(host=_OLLAMA_HOST)
    model_names = [m["model"] for m in client.list().get("models", [])]
    for candidate in _GPT_MODEL_CANDIDATES:
        if any(candidate in m for m in model_names):
            return client, candidate
    raise RuntimeError(f"사용 가능한 Ollama 모델 없음. 설치된 모델: {model_names}")


def _extract_json_block(raw: str) -> str:
    if "```json" in raw:
        return raw.split("```json")[1].split("```")[0].strip()
    if "```" in raw:
        return raw.split("```")[1].split("```")[0].strip()
    return raw.strip()


def build_llm_diagnosis_prompt(
    scorecard: Dict[str, object],
    extra: Optional[Dict[str, object]] = None,
) -> str:
    """
    Ollama(gpt-oss)에게 전달할 user prompt를 생성한다.
    - 매출 scorecard + extra에 포함된 이상징후를 기반으로 전사 매출 알림 문구를 생성한다.
    - extra에 주문수/객단가/채널/매장TOP 데이터가 있으면 현실적인 원인과 제안을 작성한다.
    - 없는 데이터는 원인처럼 단정하지 않고, 확인 필요 항목으로만 언급한다.
    """
    extra = extra or {}
    target_date = scorecard.get("target_date")
    grade = scorecard.get("grade")
    total_score = scorecard.get("total_score")

    # 파생 지표: 이달 일평균 vs 어제 비교 (LLM이 해석에 직접 활용)
    mtd_curr = scorecard.get("mtd_curr") or 0
    today_sales = scorecard.get("today_sales") or 0
    elapsed_days = 0
    try:
        from datetime import datetime as _dt
        elapsed_days = _dt.strptime(str(target_date), "%Y-%m-%d").day
    except Exception:
        pass
    mtd_daily_avg = round(mtd_curr / elapsed_days) if elapsed_days > 0 else None
    today_vs_mtd_avg_pct = (
        round((today_sales - mtd_daily_avg) / mtd_daily_avg * 100, 1)
        if mtd_daily_avg
        else None
    )

    payload = {
        "기준일": target_date,
        "등급": grade,
        "종합점수": f"{total_score}점 / 8점",
        "단기변화율": scorecard.get("short_pct"),
        "중기변화율": scorecard.get("mid_pct"),
        "장기변화율": scorecard.get("long_pct"),
        "전년변화율": scorecard.get("yoy_pct"),
        "단기점수": scorecard.get("short_score"),
        "중기점수": scorecard.get("mid_score"),
        "장기점수": scorecard.get("long_score"),
        "전년점수": scorecard.get("yoy_score"),
        "어제매출": today_sales,
        "단기기준(같은요일4주평균)": scorecard.get("short_base_avg"),
        "최근14일누적": scorecard.get("mid_recent14"),
        "이전14일누적": scorecard.get("mid_prev14"),
        "이번달누적(MTD)": mtd_curr,
        "전월동일누적(MTD)": scorecard.get("long_prev_mtd"),
        "전년동월동일누적(MTD)": scorecard.get("yoy_prev_mtd"),
        "이달경과일수": elapsed_days,
        "이달일평균(MTD/경과일수)": mtd_daily_avg,
        "어제_vs_이달일평균": f"{today_vs_mtd_avg_pct:+.1f}%" if today_vs_mtd_avg_pct is not None else None,
    }
    payload.update(extra)

    return f"""
너는 프랜차이즈 전사 매출 알림의 AI 진단 문구를 작성하는 운영 분석가다.

목표:
- 매출 상태를 짧고 명확하게 요약한다.
- 단순한 일반론이 아니라, 입력 데이터에서 확인 가능한 현실적인 원인 후보를 제시한다.
- 담당자가 바로 확인할 수 있는 구체적인 다음 액션을 제안한다.

출력 규칙:
1) 출력은 반드시 JSON 1개만 출력한다. JSON 외 다른 텍스트는 금지한다.
2) JSON 키는 정확히 다음 3개만 사용한다: situation, cause, suggest
3) 각 값은 반드시 1줄 문장으로 작성한다.
4) 각 문장은 너무 길지 않게 작성한다.
5) 숫자는 반드시 포함한다.
6) 원인은 단정하지 말고 '가능성이 있습니다', '우선 의심됩니다', '확인 필요합니다' 중 하나를 사용한다.
7) 입력 데이터에 없는 사실은 절대 만들지 않는다.
8) suggest는 뻔한 말이 아니라, 입력 데이터에 근거한 우선 확인 순서를 제안한다

situation 작성 규칙:
- situation은 반드시 아래 형식으로 작성한다.
- "어제 매출은 {{어제매출}}원이고, 같은요일 4주 평균 {{단기기준(같은요일4주평균)}}원 대비 {{단기변화율}}입니다. 이번달 누적은 {{이번달누적(MTD)}}원이고, 전월 동일기간 누적 {{전월동일누적(MTD)}}원 대비 {{장기변화율}}입니다."
- 어제매출, 단기기준, 이번달누적, 전월동일누적은 반드시 원 단위 금액으로 표시한다.
- 단기변화율은 "어제보다"가 아니라 "같은요일 4주 평균 대비"로 표현한다.
- 종합점수는 situation에 넣지 않는다.

중요한 판단 기준:
- 현재 보유 데이터는 일자별 매장별 주문건수와 매출뿐이다.
- 채널별 매출, 광고, 노출수, 클릭수, 전환율, 배민, 쿠팡 데이터는 없다.
- 따라서 채널/광고/배민/쿠팡 원인은 절대 말하지 않는다.
- 원인은 매출 변화율, 주문건수 변화, 객단가 변화, 매장별 하락 기여도 안에서만 작성한다.
- 제공된 매장별 데이터가 있으면 하락폭이 큰 매장을 우선 언급한다.
- 제공된 주문건수/객단가 데이터가 있으면 매출 하락 원인을 주문건수 하락형인지 객단가 하락형인지 구분한다.
- 단기만 하락하고 중기/장기/전년이 상승이면 전사 위기가 아니라 단기 일매출 약세로 표현한다.
- 어제 매출은 하락했지만 MTD 또는 전년동월누적이 상승이면 일시적 하락 가능성과 추세 확인 필요로 표현한다.

0점 지표 처리 규칙 (매우 중요):
- 점수 기준: 단기 0점=0%~-4.9%, 중기/장기 0점=0%~-2.9%, 전년 0점=0%~-4.9%
- 0점 지표는 정상 범위다. cause와 suggest에서 0점 지표를 하락·감소·조사 대상으로 절대 쓰지 않는다.
- 0점 범위 수치(예: 중기 -1.47%, 중기점수=0)를 "감소세", "하락 추세", "감소 중"으로 묘사하지 않는다.
- 0점 지표끼리의 차이를 "상반된 결과", "엇갈린 신호"로 해석하지 않는다.
- 종합점수 0점(정상)이면 위험·경계 신호가 없다. "우선 의심됩니다", "가능성이 있습니다(하락 관련)" 같은 경고성 표현을 쓰지 않는다.
- 종합점수 0점이면 양호한 지표의 원동력을 설명하거나, 상승 모멘텀을 어떻게 활용할지에 집중한다.

데이터 부족 시 처리:
- 채널별 데이터가 없으면 배민/쿠팡/홀 하락을 원인으로 쓰지 않는다.
- 광고 데이터가 없으면 노출/클릭/전환율 문제를 원인으로 쓰지 않는다.
- 휴무/운영시간 데이터가 없으면 확정 원인처럼 쓰지 않는다.
- 매장별 데이터가 없으면 특정 매장명을 만들지 않는다.
- 주문건수/객단가 데이터가 없으면 해당 원인을 단정하지 않는다.

금지 표현:
- 배민 매출 감소
- 쿠팡 매출 감소
- 배달 유입 둔화
- 광고 노출 감소
- 클릭 전환 흐름
- 채널별 매출 하락
- 프로모션을 재검토하세요
- 인력 배치를 재검토하세요
- TOP5를 확인하세요
- 데이터를 추가 확인하세요
- 0점 지표를 이용한 "감소세를 보였음", "하락 추세", "상반된 결과"
- 정상(0점) 등급에서 "우선 의심됩니다", "일시적일 가능성"(상승 결과에 대한 불필요한 의심)

cause 작성 규칙 — 지표 간 차이의 비즈니스 의미를 해석하라:
- 단순 나열 금지: "단기가 상승하고 장기도 상승했으므로"처럼 숫자를 다시 읽는 문장은 작성하지 않는다.
- 반드시 두 지표의 차이 또는 모순을 해석해야 한다.
  - MTD가 단기보다 훨씬 크다 → "이달 초반 며칠이 최근보다 훨씬 강세였고, 그 성과가 월 누적을 끌어올린 것"
  - 단기만 하락, MTD는 양호 → "전사 추세 이상이 아니라, 1~2개 매장의 당일 하락이 전체를 낮춘 것"
  - 단기/중기 모두 하락 → "최근 2주 동안 하락이 누적 중이며 일시적이 아닐 가능성"
  - 단기·MTD 모두 강하게 상승 → "어제도 이달 일평균 수준을 넘어서며 성장 모멘텀이 유지되는 흐름"
  - 단기 상승·MTD 상승, 중기가 0점 범위의 소폭 마이너스 → 0점 중기는 언급하지 않고 단기·MTD의 양호함을 해석
- 입력 데이터에 `이달일평균(MTD/경과일수)`과 `어제_vs_이달일평균`이 있으면 반드시 활용한다.
  - 예: "이달 일평균(약 5,700만원)보다 어제(4,500만원)가 낮았다 → 이달 초 성과가 최근보다 강했을 가능성"
- 반드시 실제 숫자를 1개 이상 포함한다.
- 0점 지표를 cause에 언급하지 않는다. 0점 범위의 수치(예: 중기 -1.47%, 0점)를 "감소", "하락", "약세"로 표현하지 않는다.

suggest 작성 규칙 — "확인하세요"로 끝내지 마라. 반드시 조건부 판단 경로로 끝내라:
- 형식: "[첫 확인 대상]을 먼저 보면, [결과 A]라면 [다음 단계 X]를, [결과 B]라면 [다음 단계 Y]를 확인하세요."
- 하락 케이스:
  - 감소폭 큰 매장 → 주문건수 같이 빠졌다면 당일 운영 중단·조기 마감이 원인, 주문건수는 유지인데 매출만 빠졌다면 취소·할인·소용량 주문이 원인
- 상승/정상 케이스:
  - 이달 일평균 대비 어제가 낮다면 → 이달 초반에 매출이 집중됐는지, 특정 매장이 이달 성장을 주도하는지 확인
  - 어제가 이달 일평균 이상이라면 → 주문건수와 객단가 중 어느 쪽이 올라서인지 구분해 지속성 판단
  - 단기·MTD 모두 양호한 정상 케이스 → 상승을 주도한 매장 또는 요인을 파악해 지속성 판단에 활용
- 0점 지표에 대한 조사·확인을 제안하지 않는다. 0점 범위 수치(예: 중기 -1.47%)를 suggest에서 "먼저 찾아야 할 문제"로 쓰지 않는다.
- 금지: "확인하세요", "분석하세요", "점검하세요"로만 끝나는 문장

좋은 출력 예시 1 (단기 -7%, MTD +22.5% — 단기만 약세):
{{
  "situation": "어제 매출은 44,467,770원이고, 같은요일 4주 평균 대비 -7.0%입니다. 이번달 누적은 695,067,995원이고, 전월 동일기간 대비 +22.5%입니다.",
  "cause": "MTD +22.5%는 건재한데 어제만 -7.0% → 이달 누적을 끌어온 성장세가 어제 하루 특정 매장에서 끊겼을 가능성이 우선 의심됩니다.",
  "suggest": "같은요일 평균 대비 감소폭이 가장 큰 매장을 먼저 찾아, 그 매장의 주문건수도 같이 빠졌다면 당일 운영 중단을, 주문건수는 유지인데 매출만 빠졌다면 취소·할인 건수를 확인하세요."
}}

좋은 출력 예시 2 (단기 -6%, 중기 -6% — 2주 누적 약세):
{{
  "situation": "어제 매출은 41,499,485원이고, 같은요일 4주 평균 대비 -5.7%입니다. 최근 14일 누적은 전전 14일 대비 -6.1%로 단기·중기 모두 약세입니다.",
  "cause": "단기 -5.7%와 중기 -6.1%가 일치한다 → 어제만의 변동이 아니라 최근 2주 동안 매출 감소가 누적되고 있을 가능성이 있습니다.",
  "suggest": "최근 14일 중 하락이 집중된 날짜를 찾아, 특정 날에 감소가 몰렸다면 그날 매장별 주문건수를, 전체 기간에 걸쳐 고른 하락이라면 객단가 추세를 먼저 비교하세요."
}}

좋은 출력 예시 3 (단기 +1.7%, MTD +21.9%, 이달일평균 > 어제):
{{
  "situation": "어제 매출은 44,951,875원이고, 같은요일 4주 평균 대비 +1.7%입니다. 이번달 누적은 740,019,870원이고, 전월 동일기간 대비 +21.9%입니다.",
  "cause": "이달 일평균(약 56,924,605원)보다 어제(44,951,875원)가 약 -21% 낮다 → MTD 성장의 대부분이 이달 초반에 집중됐고 최근 며칠은 그 수준을 밑돌고 있을 가능성이 있습니다.",
  "suggest": "이달 일평균(약 5,700만원)을 넘긴 날과 밑돈 날을 나열해, 초반 고매출 날에 주문건수가 많았는지 객단가가 높았는지를 어제와 비교하면 성장 원인을 좁힐 수 있습니다."
}}

좋은 출력 예시 4 (단기 +8.3%, MTD +29%, 어제 ≥ 이달일평균):
{{
  "situation": "어제 매출은 52,900,020원이고, 같은요일 4주 평균 대비 +8.3%입니다. 이번달 누적은 653,153,685원이고, 전월 동일기간 대비 +29.0%입니다.",
  "cause": "이달 일평균(약 59,377,608원)에 어제(52,900,020원)가 근접 중이며, 단기·MTD 모두 양호 → 이달 내내 성장 흐름이 유지되고 있을 가능성이 있습니다.",
  "suggest": "주문건수와 객단가 중 어느 쪽이 오른 날인지 먼저 구분해, 주문건수가 늘었다면 신규 유입이 지속되는지를, 객단가가 높아졌다면 고액 메뉴 주문 패턴이 반복되는지를 확인하세요."
}}

좋은 출력 예시 5 (단기 +10%, 중기 -1.47%(0점 정상 범위), MTD +14% — 종합 0점 정상):
{{
  "situation": "어제 매출은 54,775,000원이고, 같은요일 4주 평균 대비 +10.0%입니다. 이번달 누적은 1,508,195,685원이고, 전월 동일기간 누적 1,323,374,785원 대비 +14.0%입니다.",
  "cause": "단기 +10.0%·MTD +14.0% 모두 양호하며 모든 지표 0점 정상 → 이달 일평균을 어제가 넘어섰다면 성장 흐름이 이달 말까지 이어지는 것이고, 밑돌았다면 이달 초반에 성과가 집중됐을 가능성이 있습니다.",
  "suggest": "어제 상승을 주도한 매장을 먼저 파악해, 주문건수가 늘었다면 신규 유입 지속 여부를, 객단가가 높아졌다면 고액 메뉴 패턴이 반복되는지를 확인하세요."
}}

입력 데이터(JSON):
{json.dumps(payload, ensure_ascii=False)}
""".strip()


def build_llm_ai_diagnosis(
    scorecard: Dict[str, object],
    enable_llm: bool = True,
    extra: Optional[Dict[str, object]] = None,
) -> Dict[str, str]:
    """
    Ollama로 3줄 진단 문구 생성. 실패 시 룰 기반으로 폴백.
    Returns: {"situation": "...", "cause": "...", "suggest": "...", "llm_used": "Y|N"}
    """
    if not enable_llm:
        diag = build_rule_based_ai_diagnosis(scorecard)
        return {**diag, "llm_used": "N", "prompt_text": ""}

    try:
        client, model = _get_ollama_client_for_diagnosis()
        prompt = build_llm_diagnosis_prompt(scorecard, extra=extra)
        system_prompt = (
            "너는 한국어 비즈니스 알림 작성자다. 반드시 JSON만 출력한다."
        )
        resp = client.chat(
            model=model,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": prompt},
            ],
            stream=False,
        )
        raw = resp.get("message", {}).get("content", "")
        parsed = json.loads(_extract_json_block(raw))
        situation = str(parsed.get("situation", "")).strip()
        cause = str(parsed.get("cause", "")).strip()
        suggest = str(parsed.get("suggest", "")).strip()
        if not (situation and cause and suggest):
            raise ValueError("LLM JSON missing keys")
        return {"situation": situation, "cause": cause, "suggest": suggest, "llm_used": "Y", "prompt_text": prompt}
    except Exception:
        diag = build_rule_based_ai_diagnosis(scorecard)
        return {**diag, "llm_used": "N", "prompt_text": ""}


def save_llm_diagnosis_to_db(scorecard: Dict[str, object], llm_diag: Dict[str, str]) -> dict:
    """scorecard + LLM 진단 결과를 DB에 저장 (프롬프트 업그레이드용 누적 로그).

    저장 테이블: public.ai_sales_alert_llm_log
    PK: sale_date (날짜 중복 시 기존 행 유지 — insert 중복 건은 duplicated로 집계됨)
    """
    from modules.transform.utility.db import postgre_db_save

    row = {
        "sale_date": str(scorecard.get("target_date", "")),
        "grade": scorecard.get("grade"),
        "total_score": scorecard.get("total_score"),
        "short_pct": scorecard.get("short_pct"),
        "mid_pct": scorecard.get("mid_pct"),
        "long_pct": scorecard.get("long_pct"),
        "yoy_pct": scorecard.get("yoy_pct"),
        "situation": llm_diag.get("situation", ""),
        "cause": llm_diag.get("cause", ""),
        "suggest": llm_diag.get("suggest", ""),
        "llm_used": llm_diag.get("llm_used", "N"),
        "prompt_text": llm_diag.get("prompt_text", ""),
    }
    df = pd.DataFrame([row])
    return postgre_db_save(df, table="ai_sales_alert_llm_log", pk_col="sale_date", if_exists="append")


def _load_daily_sales_from_summary_sheet(
    integrated_path: Path,
    sheet_name: str = "날짜별합계",
) -> Dict[str, int]:
    wb = openpyxl.load_workbook(integrated_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"summary sheet not found: {sheet_name} ({integrated_path})")
    ws = wb[sheet_name]
    out: Dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        d, _, sales = row[:3]
        if isinstance(d, str) and d.strip():
            out[d.strip()] = _safe_int(sales)
    return out


def _load_daily_sales_from_summary_sheet(
    integrated_path: Path,
    sheet_name: str = DAILY_SUMMARY_SHEET_NAME,
) -> Dict[str, int]:
    del sheet_name
    totals_df = load_ai_daily_collection_daily_totals(base_dir=integrated_path.parent)
    return {
        str(row.sale_date): _safe_int(row.sales)
        for row in totals_df.itertuples(index=False)
    }


def build_ai_score_sheet(
    integrated_path: Path,
    target_date: str,
    sheet_name: str = "AI_진단",
) -> Dict[str, object]:
    """
    parquet에서 scorecard를 계산해 반환한다. xlsx write 없음.
    (AI 진단은 이메일 HTML로만 전달)
    """
    del sheet_name
    daily_sales = _load_daily_sales_from_summary_sheet(integrated_path)
    scorecard = compute_ai_sales_scorecard(daily_sales=daily_sales, target_date=target_date)
    out: Dict[str, object] = dict(scorecard)
    return out


def _legacy_build_ai_sales_alert_html_1(
    scorecard_with_diag: Dict[str, object],
    dashboard_url: Optional[str] = None,
) -> str:
    """
    이메일/알림용 HTML 본문(가독성 중심)을 생성한다.
    - 수치/점수/등급 + 3줄 진단문구를 카드/표 형태로 노출
    """

    def fmt_int(value: object) -> str:
        try:
            return f"{int(value):,}"
        except Exception:
            return "-"

    def fmt_signed_int(value: object) -> str:
        try:
            n = int(value)
        except Exception:
            return "-"
        sign = "+" if n > 0 else ""
        return f"{sign}{n:,}"

    def fmt_pct(value: object) -> str:
        if value is None:
            return "N/A"
        try:
            return f"{float(value):+.1f}%"
        except Exception:
            return "N/A"

    def color_if_negative(text: str) -> str:
        t = (text or "").strip()
        if t.startswith("-"):
            return f'<span style="color:#e74c3c; font-weight:700;">{t}</span>'
        return t

    target_date = str(scorecard_with_diag.get("target_date") or "-")
    grade = str(scorecard_with_diag.get("grade") or "-")
    total_score = scorecard_with_diag.get("total_score")

    today_sales = fmt_int(scorecard_with_diag.get("today_sales"))
    short_base_avg = fmt_int(scorecard_with_diag.get("short_base_avg"))

    _today_sales_n = int(scorecard_with_diag.get("today_sales") or 0)
    _short_base_avg_n = int(scorecard_with_diag.get("short_base_avg") or 0)
    short_diff = fmt_signed_int(_today_sales_n - _short_base_avg_n)

    _mid_recent14_n = int(scorecard_with_diag.get("mid_recent14") or 0)
    _mid_prev14_n = int(scorecard_with_diag.get("mid_prev14") or 0)
    mid_recent14 = fmt_int(_mid_recent14_n)
    mid_prev14 = fmt_int(_mid_prev14_n)
    mid_diff = fmt_signed_int(_mid_recent14_n - _mid_prev14_n)

    _mtd_curr_n = int(scorecard_with_diag.get("mtd_curr") or 0)
    _long_prev_mtd_n = int(scorecard_with_diag.get("long_prev_mtd") or 0)
    mtd_curr = fmt_int(_mtd_curr_n)
    long_prev_mtd = fmt_int(_long_prev_mtd_n)
    long_diff = fmt_signed_int(_mtd_curr_n - _long_prev_mtd_n)

    _yoy_prev_mtd_n = int(scorecard_with_diag.get("yoy_prev_mtd") or 0)
    yoy_prev_mtd = fmt_int(_yoy_prev_mtd_n)
    yoy_diff = fmt_signed_int(_mtd_curr_n - _yoy_prev_mtd_n)

    short_pct = fmt_pct(scorecard_with_diag.get("short_pct"))
    mid_pct = fmt_pct(scorecard_with_diag.get("mid_pct"))
    long_pct = fmt_pct(scorecard_with_diag.get("long_pct"))
    yoy_pct = fmt_pct(scorecard_with_diag.get("yoy_pct"))

    short_score = fmt_int(scorecard_with_diag.get("short_score"))
    mid_score = fmt_int(scorecard_with_diag.get("mid_score"))
    long_score = fmt_int(scorecard_with_diag.get("long_score"))
    yoy_score = fmt_int(scorecard_with_diag.get("yoy_score"))

    situation = str(scorecard_with_diag.get("situation") or "").lstrip("- ").strip()
    cause = str(scorecard_with_diag.get("cause") or "").lstrip("- ").strip()
    suggest = str(scorecard_with_diag.get("suggest") or "").lstrip("- ").strip()

    badge_color = {
        "정상": "#2ecc71",
        "주의": "#f1c40f",
        "경고": "#e67e22",
        "위험": "#e74c3c",
    }.get(grade, "#34495e")

    link_html = ""
    if dashboard_url:
        link_html = f"""
        <div style="margin-top:12px;">
          <a href="{dashboard_url}" style="color:#2c7be5; text-decoration:none;">대시보드 링크</a>
        </div>
        """

    return f"""<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="font-family:'Malgun Gothic', Arial, sans-serif; background:#f6f8fb; margin:0; padding:18px;">
  <div style="max-width:860px; margin:0 auto;">
    <div style="background:#ffffff; border:1px solid #e6e9ef; border-radius:12px; padding:18px 18px 14px 18px;">
      <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
        <div style="font-size:18px; font-weight:700; color:#111827;">[AI 진단] 전사 매출 알림</div>
        <div style="font-size:12px; color:#6b7280;">기준일: <b>{target_date}</b></div>
        <div style="margin-left:auto; display:flex; align-items:center; gap:8px;">
          <div style="font-size:12px; color:#6b7280;">종합점수: <b>{total_score}</b>/8</div>
          <div style="font-size:12px; font-weight:700; color:#ffffff; background:{badge_color}; padding:4px 10px; border-radius:999px;">{grade}</div>
        </div>
      </div>

      <div style="margin-top:14px; display:grid; grid-template-columns: 1fr; gap:10px;">
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">1. 현상황</div>
          <div style="color:#111827; line-height:1.6;">{situation}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">2. 예상 원인</div>
          <div style="color:#111827; line-height:1.6;">{cause}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">3. 제안방향</div>
          <div style="color:#111827; line-height:1.6;">{suggest}</div>
        </div>
      </div>

      <div style="margin-top:14px; padding:12px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px;">
        <div style="font-weight:700; margin-bottom:10px;">핵심 지표</div>
        <div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">
        <table cellpadding="0" cellspacing="0" style="width:100%; border-collapse:collapse; font-size:12px; table-layout:fixed;">
          <tr>
            <th style="text-align:left; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">구분</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">증감</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">변화율</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">점수</th>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">단기(어제 vs 4주 같은요일 평균)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(short_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(short_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{short_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">중기(최근 14일 누적 vs 이전 14일)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(mid_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(mid_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{mid_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">장기(MTD vs 전월 동일 경과일)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(long_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(long_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{long_score}</td>
          </tr>
          <tr>
            <td style="padding:10px;">전년(MTD vs 전년 동월 동일 경과일)</td>
            <td style="padding:10px; text-align:right;">{color_if_negative(yoy_diff)}</td>
            <td style="padding:10px; text-align:right; font-weight:700;">{color_if_negative(yoy_pct)}</td>
            <td style="padding:10px; text-align:right;">{yoy_score}</td>
          </tr>
        </table>
        </div>

        <div style="margin-top:12px; background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:10px 12px;">
          <div style="font-weight:700; color:#374151; margin-bottom:8px;">점수/등급 기준(주의 포함)</div>
          <div style="font-size:12px; color:#374151; line-height:1.6;">
            <table cellpadding="0" cellspacing="0" style="width:100%; border-collapse:collapse; font-size:12px;">
              <tr>
                <th style="text-align:left; padding:8px; border-bottom:1px solid #eef2f7; color:#6b7280;">기준</th>
                <th style="text-align:left; padding:8px; border-bottom:1px solid #eef2f7; color:#6b7280;">0점</th>
                <th style="text-align:left; padding:8px; border-bottom:1px solid #eef2f7; color:#6b7280;">1점(주의)</th>
                <th style="text-align:left; padding:8px; border-bottom:1px solid #eef2f7; color:#6b7280;">2점</th>
              </tr>
              <tr>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">단기</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">0% 이상</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">-5% ~ -10%</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">-10% 이하</td>
              </tr>
              <tr>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">중기</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">0% 이상</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">-3% ~ -7%</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">-7% 이하</td>
              </tr>
              <tr>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">장기</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">0% 이상</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">-3% ~ -7%</td>
                <td style="padding:8px; border-bottom:1px solid #f3f4f6;">-7% 이하</td>
              </tr>
              <tr>
                <td style="padding:8px;">전년</td>
                <td style="padding:8px;">0% 이상</td>
                <td style="padding:8px;">-5% ~ -10%</td>
                <td style="padding:8px;">-10% 이하</td>
              </tr>
            </table>
            <div style="margin-top:10px;">종합 등급: 0점=정상 / 1~2점=주의 / 3~4점=경고 / 5~8점=위험</div>
          </div>
        </div>
        {link_html}
      </div>

      <div style="margin-top:12px; color:#9ca3af; font-size:11px;">
        본 메일은 Airflow에서 자동 발송되었습니다.
      </div>
    </div>
  </div>
</body>
</html>"""


def _legacy_build_ai_sales_alert_html_2(
    scorecard_with_diag: Dict[str, object],
    dashboard_url: Optional[str] = None,
) -> str:
    def fmt_int(value: object) -> str:
        try:
            return f"{int(value):,}"
        except Exception:
            return "-"

    def fmt_signed_int(value: object) -> str:
        try:
            n = int(value)
        except Exception:
            return "-"
        sign = "+" if n > 0 else ""
        return f"{sign}{n:,}"

    def fmt_pct(value: object) -> str:
        if value is None:
            return "N/A"
        try:
            return f"{float(value):+.1f}%"
        except Exception:
            return "N/A"

    def color_if_negative(text: str) -> str:
        t = (text or "").strip()
        if t.startswith("-"):
            return f'<span style="color:#e74c3c; font-weight:700;">{t}</span>'
        return t

    target_date = str(scorecard_with_diag.get("target_date") or "-")
    grade = str(scorecard_with_diag.get("grade") or "-")
    total_score = scorecard_with_diag.get("total_score")
    monthly_avg_trend = scorecard_with_diag.get("monthly_avg_trend") or []

    short_diff = fmt_signed_int(int(scorecard_with_diag.get("today_sales") or 0) - int(scorecard_with_diag.get("short_base_avg") or 0))
    mid_diff = fmt_signed_int(int(scorecard_with_diag.get("mid_recent14") or 0) - int(scorecard_with_diag.get("mid_prev14") or 0))
    long_diff = fmt_signed_int(int(scorecard_with_diag.get("mtd_curr") or 0) - int(scorecard_with_diag.get("long_prev_mtd") or 0))
    yoy_diff = fmt_signed_int(int(scorecard_with_diag.get("mtd_curr") or 0) - int(scorecard_with_diag.get("yoy_prev_mtd") or 0))

    short_pct = fmt_pct(scorecard_with_diag.get("short_pct"))
    mid_pct = fmt_pct(scorecard_with_diag.get("mid_pct"))
    long_pct = fmt_pct(scorecard_with_diag.get("long_pct"))
    yoy_pct = fmt_pct(scorecard_with_diag.get("yoy_pct"))

    short_score = fmt_int(scorecard_with_diag.get("short_score"))
    mid_score = fmt_int(scorecard_with_diag.get("mid_score"))
    long_score = fmt_int(scorecard_with_diag.get("long_score"))
    yoy_score = fmt_int(scorecard_with_diag.get("yoy_score"))

    situation = str(scorecard_with_diag.get("situation") or "").lstrip("- ").strip()
    cause = str(scorecard_with_diag.get("cause") or "").lstrip("- ").strip()
    suggest = str(scorecard_with_diag.get("suggest") or "").lstrip("- ").strip()

    badge_color = {
        "정상": "#2ecc71",
        "주의": "#f1c40f",
        "경고": "#e67e22",
        "위험": "#e74c3c",
    }.get(grade, "#34495e")

    link_html = ""
    if dashboard_url:
        link_html = f"""
        <div style="margin-top:12px;">
          <a href="{dashboard_url}" style="color:#2c7be5; text-decoration:none;">대시보드 바로가기</a>
        </div>
        """

    trend_chart_html = ""
    if monthly_avg_trend:
        chart_width = 760
        chart_height = 170
        left_pad = 56
        right_pad = 28
        top_pad = 22
        bottom_pad = 44
        plot_width = chart_width - left_pad - right_pad
        plot_height = chart_height - top_pad - bottom_pad
        min_avg = min(int(item.get("avg_sales") or 0) for item in monthly_avg_trend)
        max_avg = max(int(item.get("avg_sales") or 0) for item in monthly_avg_trend)
        span = max(max_avg - min_avg, 1)

        points = []
        for idx, item in enumerate(monthly_avg_trend):
            avg_value = int(item.get("avg_sales") or 0)
            x = left_pad if len(monthly_avg_trend) == 1 else left_pad + int(plot_width * idx / (len(monthly_avg_trend) - 1))
            y = top_pad + int((max_avg - avg_value) / span * plot_height)
            points.append((x, y, item))

        grid_lines = []
        for ratio in (0.0, 0.5, 1.0):
            y = top_pad + int(plot_height * ratio)
            value = int(round(max_avg - (span * ratio)))
            grid_lines.append(f'<line x1="{left_pad}" y1="{y}" x2="{chart_width-right_pad}" y2="{y}" stroke="#e5e7eb" stroke-width="1" stroke-dasharray="3 3" />')
            grid_lines.append(f'<text x="8" y="{y + 4}" fill="#94a3b8" font-size="11">{fmt_int(value)}</text>')

        polyline_points = " ".join(f"{x},{y}" for x, y, _ in points)
        point_marks = []
        summary_cells = []
        for x, y, item in points:
            avg_sales = fmt_int(item.get("avg_sales"))
            range_label = str(item.get("range_label") or "-")
            total_sales = fmt_int(item.get("total_sales"))
            day_count = int(item.get("day_count") or 0)
            point_marks.append(
                f"""
                <circle cx="{x}" cy="{y}" r="5" fill="#2c7be5" />
                <circle cx="{x}" cy="{y}" r="9" fill="rgba(44,123,229,0.16)" />
                <text x="{x}" y="{y - 14}" text-anchor="middle" fill="#1f2937" font-size="11" font-weight="700">{avg_sales}</text>
                <text x="{x}" y="{chart_height - 12}" text-anchor="middle" fill="#64748b" font-size="11">{range_label}</text>
                """
            )
            summary_cells.append(
                f"""
                <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:10px 12px;">
                  <div style="font-size:12px; color:#64748b;">{range_label}</div>
                  <div style="margin-top:4px; font-size:15px; font-weight:700; color:#111827;">일평균 {avg_sales}</div>
                  <div style="margin-top:2px; font-size:12px; color:#64748b;">{total_sales} / {day_count}일</div>
                </div>
                """
            )

        trend_chart_html = f"""
        <div style="margin-top:14px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; color:#111827; margin-bottom:6px;">최근 2개월 일평균 매출 추이</div>
          <div style="font-size:12px; color:#6b7280; margin-bottom:10px;">예시: 4/1~10 매출 100,000 = 100,000 / 10일</div>
          <div style="background:linear-gradient(180deg, #f8fbff 0%, #ffffff 100%); border:1px solid #e6eef9; border-radius:12px; padding:8px;">
            <svg viewBox="0 0 {chart_width} {chart_height}" width="100%" height="170" role="img" aria-label="최근 2개월 일평균 매출 라인차트">
              {''.join(grid_lines)}
              <polyline fill="none" stroke="#2c7be5" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" points="{polyline_points}" />
              {''.join(point_marks)}
            </svg>
          </div>
          <div style="display:grid; grid-template-columns:repeat(2, minmax(0, 1fr)); gap:10px; margin-top:10px;">
            {''.join(summary_cells)}
          </div>
        </div>
        """

    return f"""<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="font-family:'Malgun Gothic', Arial, sans-serif; background:#f6f8fb; margin:0; padding:18px;">
  <div style="max-width:860px; margin:0 auto;">
    <div style="background:#ffffff; border:1px solid #e6e9ef; border-radius:12px; padding:18px 18px 14px 18px;">
      <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
        <div style="font-size:18px; font-weight:700; color:#111827;">[AI 진단] 본사 매출 알림</div>
        <div style="font-size:12px; color:#6b7280;">기준일 <b>{target_date}</b></div>
        <div style="margin-left:auto; display:flex; align-items:center; gap:8px;">
          <div style="font-size:12px; color:#6b7280;">종합점수: <b>{total_score}</b>/8</div>
          <div style="font-size:12px; font-weight:700; color:#ffffff; background:{badge_color}; padding:4px 10px; border-radius:999px;">{grade}</div>
        </div>
      </div>

      {trend_chart_html}

      <div style="margin-top:14px; display:grid; grid-template-columns: 1fr; gap:10px;">
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">1. 현상</div>
          <div style="color:#111827; line-height:1.6;">{situation}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">2. 예상 원인</div>
          <div style="color:#111827; line-height:1.6;">{cause}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">3. 제안 방향</div>
          <div style="color:#111827; line-height:1.6;">{suggest}</div>
        </div>
      </div>

      <div style="margin-top:14px; padding:12px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px;">
        <div style="font-weight:700; margin-bottom:10px;">상세 지표</div>
        <div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">
        <table cellpadding="0" cellspacing="0" style="width:100%; border-collapse:collapse; font-size:12px; table-layout:fixed;">
          <tr>
            <th style="text-align:left; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">구분</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">증감</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">변화율</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">점수</th>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">단기(어제 vs 4주 동일요일 평균)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(short_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(short_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{short_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">중기(최근 14일 실적 vs 이전 14일)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(mid_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(mid_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{mid_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">장기(MTD vs 전월 동일 경과)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(long_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(long_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{long_score}</td>
          </tr>
          <tr>
            <td style="padding:10px;">전년(MTD vs 전년 동월 동일 경과)</td>
            <td style="padding:10px; text-align:right;">{color_if_negative(yoy_diff)}</td>
            <td style="padding:10px; text-align:right; font-weight:700;">{color_if_negative(yoy_pct)}</td>
            <td style="padding:10px; text-align:right;">{yoy_score}</td>
          </tr>
        </table>
        </div>

        <div style="margin-top:12px; background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:10px 12px;">
          <div style="font-weight:700; color:#374151; margin-bottom:8px;">점수/등급 기준</div>
          <div style="font-size:12px; color:#374151; line-height:1.6;">
            단기: 0점 0% 이상 / 1점 -5%~-10% / 2점 -10% 이하<br/>
            중기: 0점 0% 이상 / 1점 -3%~-7% / 2점 -7% 이하<br/>
            장기: 0점 0% 이상 / 1점 -3%~-7% / 2점 -7% 이하<br/>
            전년: 0점 0% 이상 / 1점 -5%~-10% / 2점 -10% 이하<br/>
            종합 등급: 0점 정상 / 1~2점 주의 / 3~4점 경고 / 5~8점 위험
          </div>
        </div>
        {link_html}
      </div>

      <div style="margin-top:12px; color:#9ca3af; font-size:11px;">
        본 메일은 Airflow에서 자동 발송되었습니다.
      </div>
    </div>
  </div>
</body>
</html>"""


def _legacy_build_ai_sales_alert_html_3(
    scorecard_with_diag: Dict[str, object],
    dashboard_url: Optional[str] = None,
) -> str:
    def fmt_int(value: object) -> str:
        try:
            return f"{int(value):,}"
        except Exception:
            return "-"

    def fmt_signed_int(value: object) -> str:
        try:
            n = int(value)
        except Exception:
            return "-"
        sign = "+" if n > 0 else ""
        return f"{sign}{n:,}"

    def fmt_pct(value: object) -> str:
        if value is None:
            return "N/A"
        try:
            return f"{float(value):+.1f}%"
        except Exception:
            return "N/A"

    def color_if_negative(text: str) -> str:
        t = (text or "").strip()
        if t.startswith("-"):
            return f'<span style="color:#e74c3c; font-weight:700;">{t}</span>'
        return t

    target_date = str(scorecard_with_diag.get("target_date") or "-")
    grade = str(scorecard_with_diag.get("grade") or "-")
    total_score = scorecard_with_diag.get("total_score")
    monthly_avg_trend = scorecard_with_diag.get("monthly_avg_trend") or []

    short_diff = fmt_signed_int(
        int(scorecard_with_diag.get("today_sales") or 0) - int(scorecard_with_diag.get("short_base_avg") or 0)
    )
    mid_diff = fmt_signed_int(
        int(scorecard_with_diag.get("mid_recent14") or 0) - int(scorecard_with_diag.get("mid_prev14") or 0)
    )
    long_diff = fmt_signed_int(
        int(scorecard_with_diag.get("mtd_curr") or 0) - int(scorecard_with_diag.get("long_prev_mtd") or 0)
    )
    yoy_diff = fmt_signed_int(
        int(scorecard_with_diag.get("mtd_curr") or 0) - int(scorecard_with_diag.get("yoy_prev_mtd") or 0)
    )

    short_pct = fmt_pct(scorecard_with_diag.get("short_pct"))
    mid_pct = fmt_pct(scorecard_with_diag.get("mid_pct"))
    long_pct = fmt_pct(scorecard_with_diag.get("long_pct"))
    yoy_pct = fmt_pct(scorecard_with_diag.get("yoy_pct"))

    short_score = fmt_int(scorecard_with_diag.get("short_score"))
    mid_score = fmt_int(scorecard_with_diag.get("mid_score"))
    long_score = fmt_int(scorecard_with_diag.get("long_score"))
    yoy_score = fmt_int(scorecard_with_diag.get("yoy_score"))

    situation = str(scorecard_with_diag.get("situation") or "").lstrip("- ").strip()
    cause = str(scorecard_with_diag.get("cause") or "").lstrip("- ").strip()
    suggest = str(scorecard_with_diag.get("suggest") or "").lstrip("- ").strip()

    badge_color = {
        "정상": "#2ecc71",
        "주의": "#f1c40f",
        "경고": "#e67e22",
        "위험": "#e74c3c",
    }.get(grade, "#34495e")

    link_html = ""
    if dashboard_url:
        link_html = f"""
        <div style="margin-top:12px;">
          <a href="{dashboard_url}" style="color:#2c7be5; text-decoration:none;">대시보드 바로가기</a>
        </div>
        """

    trend_chart_html = ""
    if monthly_avg_trend:
        chart_width = 760
        chart_height = 170
        left_pad = 56
        right_pad = 28
        top_pad = 22
        bottom_pad = 44
        plot_width = chart_width - left_pad - right_pad
        plot_height = chart_height - top_pad - bottom_pad

        min_avg = min(int(item.get("avg_sales") or 0) for item in monthly_avg_trend)
        max_avg = max(int(item.get("avg_sales") or 0) for item in monthly_avg_trend)
        span = max(max_avg - min_avg, 1)

        points = []
        for idx, item in enumerate(monthly_avg_trend):
            avg_value = int(item.get("avg_sales") or 0)
            x = left_pad if len(monthly_avg_trend) == 1 else left_pad + int(plot_width * idx / (len(monthly_avg_trend) - 1))
            y = top_pad + int((max_avg - avg_value) / span * plot_height)
            points.append((x, y, item))

        grid_lines = []
        for ratio in (0.0, 0.5, 1.0):
            y = top_pad + int(plot_height * ratio)
            value = int(round(max_avg - (span * ratio)))
            grid_lines.append(
                f'<line x1="{left_pad}" y1="{y}" x2="{chart_width-right_pad}" y2="{y}" stroke="#e5e7eb" stroke-width="1" stroke-dasharray="3 3" />'
            )
            grid_lines.append(
                f'<text x="8" y="{y + 4}" fill="#94a3b8" font-size="11">{fmt_int(value)}</text>'
            )

        polyline_points = " ".join(f"{x},{y}" for x, y, _ in points)
        point_marks = []
        summary_cells = []
        for x, y, item in points:
            avg_sales = fmt_int(item.get("avg_sales"))
            range_label = str(item.get("range_label") or "-")
            total_sales = fmt_int(item.get("total_sales"))
            day_count = int(item.get("day_count") or 0)
            point_marks.append(
                f"""
                <circle cx="{x}" cy="{y}" r="5" fill="#2c7be5" />
                <circle cx="{x}" cy="{y}" r="9" fill="rgba(44,123,229,0.16)" />
                <text x="{x}" y="{y - 14}" text-anchor="middle" fill="#1f2937" font-size="11" font-weight="700">{avg_sales}</text>
                <text x="{x}" y="{chart_height - 12}" text-anchor="middle" fill="#64748b" font-size="11">{range_label}</text>
                """
            )
            summary_cells.append(
                f"""
                <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:10px 12px;">
                  <div style="font-size:12px; color:#64748b;">{range_label}</div>
                  <div style="margin-top:4px; font-size:15px; font-weight:700; color:#111827;">일평균 {avg_sales}</div>
                  <div style="margin-top:2px; font-size:12px; color:#64748b;">{total_sales} / {day_count}일</div>
                </div>
                """
            )

        trend_chart_html = f"""
        <div style="margin-top:14px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; color:#111827; margin-bottom:6px;">최근 2개월 일평균 매출 추이</div>
          <div style="font-size:12px; color:#6b7280; margin-bottom:10px;">예시: 4/1~10 매출 100,000 = 100,000 / 10일</div>
          <div style="background:linear-gradient(180deg, #f8fbff 0%, #ffffff 100%); border:1px solid #e6eef9; border-radius:12px; padding:8px;">
            <svg viewBox="0 0 {chart_width} {chart_height}" width="100%" height="170" role="img" aria-label="최근 2개월 일평균 매출 라인차트">
              {''.join(grid_lines)}
              <polyline fill="none" stroke="#2c7be5" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" points="{polyline_points}" />
              {''.join(point_marks)}
            </svg>
          </div>
          <div style="display:grid; grid-template-columns:repeat(2, minmax(0, 1fr)); gap:10px; margin-top:10px;">
            {''.join(summary_cells)}
          </div>
        </div>
        """

    return f"""<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="font-family:'Malgun Gothic', Arial, sans-serif; background:#f6f8fb; margin:0; padding:18px;">
  <div style="max-width:860px; margin:0 auto;">
    <div style="background:#ffffff; border:1px solid #e6e9ef; border-radius:12px; padding:18px 18px 14px 18px;">
      <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
        <div style="font-size:18px; font-weight:700; color:#111827;">[AI 진단] 본사 매출 알림</div>
        <div style="font-size:12px; color:#6b7280;">기준일 <b>{target_date}</b></div>
        <div style="margin-left:auto; display:flex; align-items:center; gap:8px;">
          <div style="font-size:12px; color:#6b7280;">종합점수: <b>{total_score}</b>/8</div>
          <div style="font-size:12px; font-weight:700; color:#ffffff; background:{badge_color}; padding:4px 10px; border-radius:999px;">{grade}</div>
        </div>
      </div>

      {trend_chart_html}

      <div style="margin-top:14px; display:grid; grid-template-columns: 1fr; gap:10px;">
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">1. 현상</div>
          <div style="color:#111827; line-height:1.6;">{situation}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">2. 예상 원인</div>
          <div style="color:#111827; line-height:1.6;">{cause}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">3. 제안 방향</div>
          <div style="color:#111827; line-height:1.6;">{suggest}</div>
        </div>
      </div>

      <div style="margin-top:14px; padding:12px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px;">
        <div style="font-weight:700; margin-bottom:10px;">상세 지표</div>
        <div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">
        <table cellpadding="0" cellspacing="0" style="width:100%; border-collapse:collapse; font-size:12px; table-layout:fixed;">
          <tr>
            <th style="text-align:left; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">구분</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">증감</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">변화율</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">점수</th>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">단기(어제 vs 4주 동일요일 평균)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(short_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(short_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{short_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">중기(최근 14일 실적 vs 이전 14일)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(mid_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(mid_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{mid_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">장기(MTD vs 전월 동일 경과)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(long_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(long_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{long_score}</td>
          </tr>
          <tr>
            <td style="padding:10px;">전년(MTD vs 전년 동월 동일 경과)</td>
            <td style="padding:10px; text-align:right;">{color_if_negative(yoy_diff)}</td>
            <td style="padding:10px; text-align:right; font-weight:700;">{color_if_negative(yoy_pct)}</td>
            <td style="padding:10px; text-align:right;">{yoy_score}</td>
          </tr>
        </table>
        </div>

        <div style="margin-top:12px; background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:10px 12px;">
          <div style="font-weight:700; color:#374151; margin-bottom:8px;">점수/등급 기준</div>
          <div style="font-size:12px; color:#374151; line-height:1.6;">
            단기: 0점 0% 이상 / 1점 -5%~-10% / 2점 -10% 이하<br/>
            중기: 0점 0% 이상 / 1점 -3%~-7% / 2점 -7% 이하<br/>
            장기: 0점 0% 이상 / 1점 -3%~-7% / 2점 -7% 이하<br/>
            전년: 0점 0% 이상 / 1점 -5%~-10% / 2점 -10% 이하<br/>
            종합 등급: 0점 정상 / 1~2점 주의 / 3~4점 경고 / 5~8점 위험
          </div>
        </div>
        {link_html}
      </div>

      <div style="margin-top:12px; color:#9ca3af; font-size:11px;">
        본 메일은 Airflow에서 자동 발송되었습니다.
      </div>
    </div>
  </div>
</body>
</html>"""


def _legacy_build_ai_sales_alert_html_4(
    scorecard_with_diag: Dict[str, object],
    dashboard_url: Optional[str] = None,
) -> str:
    def fmt_int(value: object) -> str:
        try:
            return f"{int(value):,}"
        except Exception:
            return "-"

    def fmt_signed_int(value: object) -> str:
        try:
            n = int(value)
        except Exception:
            return "-"
        sign = "+" if n > 0 else ""
        return f"{sign}{n:,}"

    def fmt_pct(value: object) -> str:
        if value is None:
            return "N/A"
        try:
            return f"{float(value):+.1f}%"
        except Exception:
            return "N/A"

    def color_if_negative(text: str) -> str:
        t = (text or "").strip()
        if t.startswith("-"):
            return f'<span style="color:#e74c3c; font-weight:700;">{t}</span>'
        return t

    target_date = str(scorecard_with_diag.get("target_date") or "-")
    grade = str(scorecard_with_diag.get("grade") or "-")
    total_score = scorecard_with_diag.get("total_score")
    monthly_avg_trend = scorecard_with_diag.get("monthly_avg_trend") or []

    today_sales_n = int(scorecard_with_diag.get("today_sales") or 0)
    short_base_avg_n = int(scorecard_with_diag.get("short_base_avg") or 0)
    short_diff = fmt_signed_int(today_sales_n - short_base_avg_n)

    mid_recent14_n = int(scorecard_with_diag.get("mid_recent14") or 0)
    mid_prev14_n = int(scorecard_with_diag.get("mid_prev14") or 0)
    mid_diff = fmt_signed_int(mid_recent14_n - mid_prev14_n)

    mtd_curr_n = int(scorecard_with_diag.get("mtd_curr") or 0)
    long_prev_mtd_n = int(scorecard_with_diag.get("long_prev_mtd") or 0)
    long_diff = fmt_signed_int(mtd_curr_n - long_prev_mtd_n)

    yoy_prev_mtd_n = int(scorecard_with_diag.get("yoy_prev_mtd") or 0)
    yoy_diff = fmt_signed_int(mtd_curr_n - yoy_prev_mtd_n)

    short_pct = fmt_pct(scorecard_with_diag.get("short_pct"))
    mid_pct = fmt_pct(scorecard_with_diag.get("mid_pct"))
    long_pct = fmt_pct(scorecard_with_diag.get("long_pct"))
    yoy_pct = fmt_pct(scorecard_with_diag.get("yoy_pct"))

    short_score = fmt_int(scorecard_with_diag.get("short_score"))
    mid_score = fmt_int(scorecard_with_diag.get("mid_score"))
    long_score = fmt_int(scorecard_with_diag.get("long_score"))
    yoy_score = fmt_int(scorecard_with_diag.get("yoy_score"))

    situation = str(scorecard_with_diag.get("situation") or "").lstrip("- ").strip()
    cause = str(scorecard_with_diag.get("cause") or "").lstrip("- ").strip()
    suggest = str(scorecard_with_diag.get("suggest") or "").lstrip("- ").strip()

    badge_color = {
        "정상": "#2ecc71",
        "주의": "#f1c40f",
        "경고": "#e67e22",
        "위험": "#e74c3c",
    }.get(grade, "#34495e")

    link_html = ""
    if dashboard_url:
        link_html = f"""
        <div style="margin-top:12px;">
          <a href="{dashboard_url}" style="color:#2c7be5; text-decoration:none;">대시보드 바로가기</a>
        </div>
        """

    trend_chart_html = ""
    if monthly_avg_trend:
        chart_width = 760
        chart_height = 170
        left_pad = 56
        right_pad = 28
        top_pad = 22
        bottom_pad = 44
        plot_width = chart_width - left_pad - right_pad
        plot_height = chart_height - top_pad - bottom_pad

        min_avg = min(int(item.get("avg_sales") or 0) for item in monthly_avg_trend)
        max_avg = max(int(item.get("avg_sales") or 0) for item in monthly_avg_trend)
        span = max(max_avg - min_avg, 1)

        points = []
        for idx, item in enumerate(monthly_avg_trend):
            avg_value = int(item.get("avg_sales") or 0)
            x = left_pad if len(monthly_avg_trend) == 1 else left_pad + int(plot_width * idx / (len(monthly_avg_trend) - 1))
            y = top_pad + int((max_avg - avg_value) / span * plot_height) if span else top_pad + plot_height // 2
            points.append((x, y, item))

        polyline_points = " ".join(f"{x},{y}" for x, y, _ in points)
        grid_lines = []
        for ratio in (0.0, 0.5, 1.0):
            y = top_pad + int(plot_height * ratio)
            value = int(round(max_avg - (span * ratio)))
            grid_lines.append(
                f'<line x1="{left_pad}" y1="{y}" x2="{chart_width-right_pad}" y2="{y}" stroke="#e5e7eb" stroke-width="1" stroke-dasharray="3 3" />'
            )
            grid_lines.append(
                f'<text x="8" y="{y + 4}" fill="#94a3b8" font-size="11">{fmt_int(value)}</text>'
            )

        point_marks = []
        summary_cells = []
        for x, y, item in points:
            avg_sales = fmt_int(item.get("avg_sales"))
            range_label = str(item.get("range_label") or "-")
            total_sales = fmt_int(item.get("total_sales"))
            day_count = int(item.get("day_count") or 0)
            point_marks.append(
                f"""
                <circle cx="{x}" cy="{y}" r="5" fill="#2c7be5" />
                <circle cx="{x}" cy="{y}" r="9" fill="rgba(44,123,229,0.16)" />
                <text x="{x}" y="{y - 14}" text-anchor="middle" fill="#1f2937" font-size="11" font-weight="700">{avg_sales}</text>
                <text x="{x}" y="{chart_height - 12}" text-anchor="middle" fill="#64748b" font-size="11">{range_label}</text>
                """
            )
            summary_cells.append(
                f"""
                <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:10px 12px;">
                  <div style="font-size:12px; color:#64748b;">{range_label}</div>
                  <div style="margin-top:4px; font-size:15px; font-weight:700; color:#111827;">일평균 {avg_sales}</div>
                  <div style="margin-top:2px; font-size:12px; color:#64748b;">{total_sales} / {day_count}일</div>
                </div>
                """
            )

        trend_chart_html = f"""
        <div style="margin-top:14px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; color:#111827; margin-bottom:6px;">최근 2개월 일평균 매출 추이</div>
          <div style="font-size:12px; color:#6b7280; margin-bottom:10px;">예시: 4/1~10 매출 100,000 = 100,000 / 10일</div>
          <div style="background:linear-gradient(180deg, #f8fbff 0%, #ffffff 100%); border:1px solid #e6eef9; border-radius:12px; padding:8px;">
            <svg viewBox="0 0 {chart_width} {chart_height}" width="100%" height="170" role="img" aria-label="최근 2개월 일평균 매출 라인차트">
              {''.join(grid_lines)}
              <polyline fill="none" stroke="#2c7be5" stroke-width="3" stroke-linecap="round" stroke-linejoin="round" points="{polyline_points}" />
              {''.join(point_marks)}
            </svg>
          </div>
          <div style="display:grid; grid-template-columns:repeat(2, minmax(0, 1fr)); gap:10px; margin-top:10px;">
            {''.join(summary_cells)}
          </div>
        </div>
        """

    return f"""<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="font-family:'Malgun Gothic', Arial, sans-serif; background:#f6f8fb; margin:0; padding:18px;">
  <div style="max-width:860px; margin:0 auto;">
    <div style="background:#ffffff; border:1px solid #e6e9ef; border-radius:12px; padding:18px 18px 14px 18px;">
      <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
        <div style="font-size:18px; font-weight:700; color:#111827;">[AI 진단] 본사 매출 알림</div>
        <div style="font-size:12px; color:#6b7280;">기준일 <b>{target_date}</b></div>
        <div style="margin-left:auto; display:flex; align-items:center; gap:8px;">
          <div style="font-size:12px; color:#6b7280;">종합점수: <b>{total_score}</b>/8</div>
          <div style="font-size:12px; font-weight:700; color:#ffffff; background:{badge_color}; padding:4px 10px; border-radius:999px;">{grade}</div>
        </div>
      </div>

      {trend_chart_html}

      <div style="margin-top:14px; display:grid; grid-template-columns: 1fr; gap:10px;">
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">1. 현상</div>
          <div style="color:#111827; line-height:1.6;">{situation}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">2. 예상 원인</div>
          <div style="color:#111827; line-height:1.6;">{cause}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">3. 제안 방향</div>
          <div style="color:#111827; line-height:1.6;">{suggest}</div>
        </div>
      </div>

      <div style="margin-top:14px; padding:12px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px;">
        <div style="font-weight:700; margin-bottom:10px;">상세 지표</div>
        <div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">
        <table cellpadding="0" cellspacing="0" style="width:100%; border-collapse:collapse; font-size:12px; table-layout:fixed;">
          <tr>
            <th style="text-align:left; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">구분</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">증감</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">변화율</th>
            <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">점수</th>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">단기(어제 vs 4주 동일요일 평균)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(short_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(short_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{short_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">중기(최근 14일 실적 vs 이전 14일)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(mid_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(mid_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{mid_score}</td>
          </tr>
          <tr>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6;">장기(MTD vs 전월 동일 경과)</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{color_if_negative(long_diff)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{color_if_negative(long_pct)}</td>
            <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{long_score}</td>
          </tr>
          <tr>
            <td style="padding:10px;">전년(MTD vs 전년 동월 동일 경과)</td>
            <td style="padding:10px; text-align:right;">{color_if_negative(yoy_diff)}</td>
            <td style="padding:10px; text-align:right; font-weight:700;">{color_if_negative(yoy_pct)}</td>
            <td style="padding:10px; text-align:right;">{yoy_score}</td>
          </tr>
        </table>
        </div>

        <div style="margin-top:12px; background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:10px 12px;">
          <div style="font-weight:700; color:#374151; margin-bottom:8px;">점수/등급 기준</div>
          <div style="font-size:12px; color:#374151; line-height:1.6;">
            단기: 0점 0% 이상 / 1점 -5%~-10% / 2점 -10% 이하<br/>
            중기: 0점 0% 이상 / 1점 -3%~-7% / 2점 -7% 이하<br/>
            장기: 0점 0% 이상 / 1점 -3%~-7% / 2점 -7% 이하<br/>
            전년: 0점 0% 이상 / 1점 -5%~-10% / 2점 -10% 이하<br/>
            종합 등급: 0점 정상 / 1~2점 주의 / 3~4점 경고 / 5~8점 위험
          </div>
        </div>
        {link_html}
      </div>

      <div style="margin-top:12px; color:#9ca3af; font-size:11px;">
        본 메일은 Airflow에서 자동 발송되었습니다.
      </div>
    </div>
  </div>
</body>
</html>"""



AI_SALES_ALERT_TREND_CHART_CID = "ai-daily-collection-trend-chart"


def _mail_fmt_int(value: object) -> str:
    try:
        return f"{int(value):,}"
    except Exception:
        return "-"


def _mail_fmt_signed_int(value: object) -> str:
    try:
        number = int(value)
    except Exception:
        return "-"
    sign = "+" if number > 0 else ""
    return f"{sign}{number:,}"


def _mail_fmt_pct(value: object) -> str:
    if value is None:
        return "N/A"
    try:
        return f"{float(value):+.1f}%"
    except Exception:
        return "N/A"


def _mail_negative_html(text: str) -> str:
    normalized = (text or "").strip()
    escaped = escape(normalized)
    if normalized.startswith("-"):
        return f'<span style="color:#dc2626; font-weight:700;">{escaped}</span>'
    return escaped


def _normalize_daily_avg_trend(daily_trend: object) -> List[Dict[str, object]]:
    normalized: List[Dict[str, object]] = []
    for item in daily_trend or []:
        if not isinstance(item, dict):
            continue
        normalized.append(
            {
                "sale_date": str(item.get("sale_date") or ""),
                "label": str(item.get("label") or "-"),
                "range_label": str(item.get("range_label") or "-"),
                "total_sales": int(item.get("total_sales") or 0),
                "day_count": int(item.get("day_count") or 0),
                "avg_sales": int(item.get("avg_sales") or 0),
                "avg_growth_pct_vs_prev_day": item.get("avg_growth_pct_vs_prev_day"),
            }
        )
    return normalized


def _load_mail_chart_font(size: int):
    from PIL import ImageFont

    font_candidates = [
        "malgun.ttf",
        "DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    ]
    for candidate in font_candidates:
        try:
            return ImageFont.truetype(candidate, size=size)
        except Exception:
            continue
    return ImageFont.load_default()


def _draw_centered_text(draw, xy, text, font, fill):
    bbox = draw.textbbox((0, 0), text, font=font)
    x = int(xy[0] - (bbox[2] - bbox[0]) / 2)
    y = int(xy[1] - (bbox[3] - bbox[1]) / 2)
    draw.text((x, y), text, font=font, fill=fill)


def build_ai_sales_alert_trend_chart_png(
    daily_trend: List[Dict[str, object]],
    width: int = 760,
    height: int = 260,
) -> bytes:
    from io import BytesIO
    from PIL import Image, ImageDraw

    trend = _normalize_daily_avg_trend(daily_trend)
    if not trend:
        return b""

    image = Image.new("RGB", (width, height), "#ffffff")
    draw = ImageDraw.Draw(image)

    value_font = _load_mail_chart_font(15)
    axis_font = _load_mail_chart_font(12)

    draw.rounded_rectangle((0, 0, width - 1, height - 1), radius=18, outline="#e5edf7", width=1, fill="#ffffff")

    left_pad = 76
    right_pad = 28
    top_pad = 26
    bottom_pad = 50
    plot_width = width - left_pad - right_pad
    plot_height = height - top_pad - bottom_pad

    min_avg = min(item["avg_sales"] for item in trend)
    max_avg = max(item["avg_sales"] for item in trend)
    span = max(max_avg - min_avg, 1)

    points = []
    for idx, item in enumerate(trend):
        x = left_pad if len(trend) == 1 else left_pad + int(plot_width * idx / (len(trend) - 1))
        y = top_pad + int((max_avg - item["avg_sales"]) / span * plot_height)
        points.append((x, y, item))

    for ratio in (0.0, 0.5, 1.0):
        y = top_pad + int(plot_height * ratio)
        value = int(round(max_avg - (span * ratio)))
        draw.line((left_pad, y, width - right_pad, y), fill="#dbe5f0", width=1)
        draw.text((12, y - 8), _mail_fmt_int(value), fill="#94a3b8", font=axis_font)

    if len(points) >= 2:
        draw.line([(x, y) for x, y, _ in points], fill="#2563eb", width=4)

    last_index = len(points) - 1
    for idx, (x, y, item) in enumerate(points):
        draw.ellipse((x - 5, y - 5, x + 5, y + 5), fill="#2563eb")
        if idx == last_index:
            draw.ellipse((x - 10, y - 10, x + 10, y + 10), outline="#bfdbfe", width=3)
            _draw_centered_text(draw, (x, y - 20), _mail_fmt_int(item["avg_sales"]), value_font, "#111827")
        if idx == 0 or idx == last_index or idx % 5 == 0:
            _draw_centered_text(draw, (x, height - 20), item["label"], axis_font, "#64748b")

    buffer = BytesIO()
    image.save(buffer, format="PNG")
    return buffer.getvalue()


def _build_ai_sales_alert_trend_html(
    scorecard_with_diag: Dict[str, object],
    daily_trend: List[Dict[str, object]],
    trend_chart_cid: Optional[str],
) -> str:
    trend = _normalize_daily_avg_trend(daily_trend)
    if not trend:
        return ""

    target_dt = pendulum.parse(str(scorecard_with_diag.get("target_date") or trend[-1]["sale_date"]))
    prev_month_dt = target_dt.subtract(months=1)
    prev_month_total = int(scorecard_with_diag.get("prev_month_total") or 0)
    prev_month_day_count = int(scorecard_with_diag.get("prev_month_day_count") or prev_month_dt.days_in_month)
    prev_month_avg = int(scorecard_with_diag.get("prev_month_avg") or 0)
    mtd_curr = int(scorecard_with_diag.get("mtd_curr") or 0)
    mtd_day_count = int(scorecard_with_diag.get("mtd_day_count") or target_dt.day)
    mtd_avg = int(scorecard_with_diag.get("mtd_avg") or 0)
    mtd_avg_vs_prev_month_avg_pct = scorecard_with_diag.get("mtd_avg_vs_prev_month_avg_pct")
    compare_color = "#64748b"
    compare_text = "\uc804\ub2ec \ube44\uad50 \uc5c6\uc74c"
    if mtd_avg_vs_prev_month_avg_pct is not None:
        compare_color = "#2563eb" if float(mtd_avg_vs_prev_month_avg_pct) >= 0 else "#dc2626"
        compare_text = f"\uc804\ub2ec \ub300\ube44 {_mail_fmt_pct(mtd_avg_vs_prev_month_avg_pct)}"

    summary_cells = [
        f"""
        <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:12px 14px;">
          <div style="font-size:12px; color:#64748b;">\uc804\ub2ec \uc77c\ud3c9\uade0</div>
          <div style="margin-top:4px; font-size:12px; color:#64748b;">{prev_month_dt.month}\uc6d4 {_mail_fmt_int(prev_month_total)} / {prev_month_day_count}</div>
          <div style="margin-top:6px; font-size:18px; font-weight:700; color:#111827;">{_mail_fmt_int(prev_month_avg)}</div>
        </div>
        """,
        f"""
        <div style="background:#f8fafc; border:1px solid #e2e8f0; border-radius:10px; padding:12px 14px;">
          <div style="font-size:12px; color:#64748b;">\uc774\ubc88\ub2ec \uc77c\ud3c9\uade0</div>
          <div style="margin-top:4px; font-size:12px; color:#64748b;">{target_dt.month}/1~{mtd_day_count} {_mail_fmt_int(mtd_curr)} / {mtd_day_count}</div>
          <div style="margin-top:6px; font-size:18px; font-weight:700; color:#111827;">{_mail_fmt_int(mtd_avg)}</div>
          <div style="margin-top:4px; font-size:12px; font-weight:700; color:{compare_color};">{escape(compare_text)}</div>
        </div>
        """,
    ]

    chart_html = ""
    if trend_chart_cid:
        chart_html = f"""
        <div style="margin-top:10px; background:linear-gradient(180deg, #f8fbff 0%, #ffffff 100%); border:1px solid #e6eef9; border-radius:12px; padding:10px;">
          <img src="cid:{escape(trend_chart_cid)}" alt="\ucd5c\uadfc 30\uc77c \uc77c\ubcc4 \uc6d4\ud3c9\uade0 \ucd94\uc774" width="760" style="display:block; width:100%; max-width:760px; height:auto; border:0;" />
        </div>
        """

    return f"""
    <div style="margin-top:14px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
      <div style="font-weight:700; color:#111827; margin-bottom:6px;">\ucd5c\uadfc 30\uc77c \uc77c\ubcc4 \uc6d4\ud3c9\uade0 \ucd94\uc774</div>
      <div style="font-size:12px; color:#6b7280;">\uac01 \uc2dc\uc810\uc758 \uc6d4 \ub204\uc801 \ub9e4\ucd9c\uc744 \ud574\ub2f9 \uc6d4 \uacbd\uacfc\uc77c\uc218\ub85c \ub098\ub208 \uac12\uc785\ub2c8\ub2e4.</div>
      {chart_html}
      <div style="display:grid; grid-template-columns:repeat(2, minmax(0, 1fr)); gap:10px; margin-top:10px;">
        {''.join(summary_cells)}
      </div>
    </div>
    """


def build_ai_sales_alert_html(
    scorecard_with_diag: Dict[str, object],
    dashboard_url: Optional[str] = None,
    trend_chart_cid: Optional[str] = None,
) -> str:
    target_date = str(scorecard_with_diag.get("target_date") or "-")
    grade = str(scorecard_with_diag.get("grade") or "-")
    total_score = scorecard_with_diag.get("total_score")
    daily_trend = _normalize_daily_avg_trend(scorecard_with_diag.get("daily_avg_trend") or [])

    short_diff = _mail_fmt_signed_int(
        int(scorecard_with_diag.get("today_sales") or 0) - int(scorecard_with_diag.get("short_base_avg") or 0)
    )
    mid_diff = _mail_fmt_signed_int(
        int(scorecard_with_diag.get("mid_recent14") or 0) - int(scorecard_with_diag.get("mid_prev14") or 0)
    )
    long_diff = _mail_fmt_signed_int(
        int(scorecard_with_diag.get("mtd_curr") or 0) - int(scorecard_with_diag.get("long_prev_mtd") or 0)
    )
    yoy_diff = _mail_fmt_signed_int(
        int(scorecard_with_diag.get("mtd_curr") or 0) - int(scorecard_with_diag.get("yoy_prev_mtd") or 0)
    )

    short_pct = _mail_fmt_pct(scorecard_with_diag.get("short_pct"))
    mid_pct = _mail_fmt_pct(scorecard_with_diag.get("mid_pct"))
    long_pct = _mail_fmt_pct(scorecard_with_diag.get("long_pct"))
    yoy_pct = _mail_fmt_pct(scorecard_with_diag.get("yoy_pct"))

    short_score = _mail_fmt_int(scorecard_with_diag.get("short_score"))
    mid_score = _mail_fmt_int(scorecard_with_diag.get("mid_score"))
    long_score = _mail_fmt_int(scorecard_with_diag.get("long_score"))
    yoy_score = _mail_fmt_int(scorecard_with_diag.get("yoy_score"))

    situation = escape(str(scorecard_with_diag.get("situation") or "").lstrip("- ").strip().replace("최근 14일", "직전 14일"))
    cause = escape(str(scorecard_with_diag.get("cause") or "").lstrip("- ").strip().replace("최근 14일", "직전 14일"))
    suggest = escape(str(scorecard_with_diag.get("suggest") or "").lstrip("- ").strip().replace("최근 14일", "직전 14일"))

    badge_color = {
        "\uc815\uc0c1": "#16a34a",
        "\uc8fc\uc758": "#d97706",
        "\uacbd\uace0": "#ea580c",
        "\uc704\ud5d8": "#dc2626",
    }.get(grade, "#475569")

    trend_html = _build_ai_sales_alert_trend_html(scorecard_with_diag, daily_trend, trend_chart_cid)
    link_html = ""
    if dashboard_url:
        link_html = (
            f'<div style="margin-top:12px;"><a href="{escape(dashboard_url, quote=True)}" '
            'style="color:#2563eb; text-decoration:none;">\ub300\uc2dc\ubcf4\ub4dc \ubc14\ub85c\uac00\uae30</a></div>'
        )

    return f"""<html>
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
</head>
<body style="font-family:'Malgun Gothic', Arial, sans-serif; background:#f6f8fb; margin:0; padding:18px;">
  <div style="max-width:860px; margin:0 auto;">
    <div style="background:#ffffff; border:1px solid #e6e9ef; border-radius:12px; padding:18px 18px 14px 18px;">
      <div style="display:flex; align-items:center; gap:10px; flex-wrap:wrap;">
        <div style="font-size:18px; font-weight:700; color:#111827;">[AI \uc9c4\ub2e8] \uc77c\ub9e4\ucd9c \uc54c\ub9bc</div>
        <div style="font-size:12px; color:#6b7280;">\uae30\uc900\uc77c: <b>{escape(target_date)}</b></div>
        <div style="margin-left:auto; display:flex; align-items:center; gap:8px;">
          <div style="font-size:12px; color:#6b7280;">\uc885\ud569\uc810\uc218: <b>{escape(str(total_score))}</b>/8</div>
          <div style="font-size:12px; font-weight:700; color:#ffffff; background:{badge_color}; padding:4px 10px; border-radius:999px;">{escape(grade)}</div>
        </div>
      </div>

      {trend_html}

      <div style="margin-top:14px; display:grid; grid-template-columns:1fr; gap:10px;">
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">1. \ud604\uc0c1</div>
          <div style="color:#111827; line-height:1.6;">{situation}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">2. \ucd94\uc815 \uc6d0\uc778</div>
          <div style="color:#111827; line-height:1.6;">{cause}</div>
        </div>
        <div style="background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:12px;">
          <div style="font-weight:700; margin-bottom:6px;">3. \uc81c\uc548 \uc561\uc158</div>
          <div style="color:#111827; line-height:1.6;">{suggest}</div>
        </div>
      </div>

      <div style="margin-top:14px; padding:12px; background:#ffffff; border:1px solid #eef2f7; border-radius:10px;">
        <div style="font-weight:700; margin-bottom:10px;">\uc0c1\uc138 \uc9c0\ud45c</div>
        <div style="overflow-x:auto; -webkit-overflow-scrolling:touch;">
          <table cellpadding="0" cellspacing="0" style="width:100%; border-collapse:collapse; font-size:12px; table-layout:fixed;">
            <tr>
              <th style="text-align:left; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">\uad6c\ubd84</th>
              <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">\uc99d\uac10</th>
              <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">\ubcc0\ud654\uc728</th>
              <th style="text-align:right; padding:10px; border-bottom:1px solid #eef2f7; color:#6b7280;">\uc810\uc218</th>
            </tr>
            <tr>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6;">\ub2e8\uae30(\uc5b4\uc81c vs 4\uc8fc \ub3d9\uc77c\uc694\uc77c \ud3c9\uade0)</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{_mail_negative_html(short_diff)}</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{_mail_negative_html(short_pct)}</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{escape(short_score)}</td>
            </tr>
            <tr>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6;">\uc911\uae30(\uc9c1\uc804 14\uc77c \ud569\uacc4 vs \uc774\uc804 14\uc77c)</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{_mail_negative_html(mid_diff)}</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{_mail_negative_html(mid_pct)}</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{escape(mid_score)}</td>
            </tr>
            <tr>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6;">\uc7a5\uae30(MTD vs \uc804\uc6d4 \ub3d9\uc77c\uae30\uac04)</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{_mail_negative_html(long_diff)}</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right; font-weight:700;">{_mail_negative_html(long_pct)}</td>
              <td style="padding:10px; border-bottom:1px solid #f3f4f6; text-align:right;">{escape(long_score)}</td>
            </tr>
            <tr>
              <td style="padding:10px;">\uc804\ub144(MTD vs \uc804\ub144 \ub3d9\uc6d4 \ub3d9\uc77c\uae30\uac04)</td>
              <td style="padding:10px; text-align:right;">{_mail_negative_html(yoy_diff)}</td>
              <td style="padding:10px; text-align:right; font-weight:700;">{_mail_negative_html(yoy_pct)}</td>
              <td style="padding:10px; text-align:right;">{escape(yoy_score)}</td>
            </tr>
          </table>
        </div>
        <div style="margin-top:12px; background:#f9fafb; border:1px solid #eef2f7; border-radius:10px; padding:10px 12px;">
          <div style="font-weight:700; color:#374151; margin-bottom:8px;">\uc810\uc218 / \ub4f1\uae09 \uae30\uc900</div>
          <div style="font-size:12px; color:#374151; line-height:1.6;">
            \ub2e8\uae30: 0\uc810 0% \uc774\uc0c1 / 1\uc810 -5%~-10% / 2\uc810 -10% \uc774\ud558<br/>
            \uc911\uae30: 0\uc810 0% \uc774\uc0c1 / 1\uc810 -3%~-7% / 2\uc810 -7% \uc774\ud558<br/>
            \uc7a5\uae30: 0\uc810 0% \uc774\uc0c1 / 1\uc810 -3%~-7% / 2\uc810 -7% \uc774\ud558<br/>
            \uc804\ub144: 0\uc810 0% \uc774\uc0c1 / 1\uc810 -5%~-10% / 2\uc810 -10% \uc774\ud558<br/>
            \uc885\ud569 \ub4f1\uae09: 0\uc810 \uc815\uc0c1 / 1~2\uc810 \uc8fc\uc758 / 3~4\uc810 \uacbd\uace0 / 5~8\uc810 \uc704\ud5d8
          </div>
        </div>
        {link_html}
      </div>
      <div style="margin-top:12px; color:#9ca3af; font-size:11px;">\ubcf8 \uba54\uc77c\uc740 Airflow\uc5d0\uc11c \uc790\ub3d9 \ubc1c\uc1a1\ub418\uc5c8\uc2b5\ub2c8\ub2e4.</div>
    </div>
  </div>
</body>
</html>"""
